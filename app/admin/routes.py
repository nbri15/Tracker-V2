"""Admin routes for school setup and management."""

from __future__ import annotations

import csv
import io
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import date, datetime, timedelta, timezone

from flask import Response, current_app, flash, jsonify, make_response, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required

from sqlalchemy import and_, func, or_
from sqlalchemy.orm import joinedload
from sqlalchemy.exc import OperationalError
from weasyprint import HTML

from app.extensions import db


def _pdf_requested() -> bool:
    return request.args.get('format', '').strip().lower() == 'pdf' or request.args.get('pdf', '0') == '1'


def _render_table_pdf(title: str, headers: list, rows: list, filters: dict | None = None, anonymised: bool = False, filename: str | None = None, subtitle: str | None = None):
    template_name = 'exports/table_pdf.html'
    safe_title = title if not anonymised else title.replace('named', 'anonymised').replace('Named', 'Anonymised')
    current_app.logger.info(
        'PDF export route hit endpoint=%s template=%s rows=%s anonymised=%s filename=%s',
        request.endpoint,
        template_name,
        len(rows),
        anonymised,
        filename or 'table_report.pdf',
    )
    try:
        html = render_template(
            template_name,
            title=safe_title,
            subtitle=subtitle,
            headers=headers,
            rows=rows,
            filters=filters or {},
            anonymise=anonymised,
            generated_at=datetime.now(timezone.utc),
        )
        pdf_bytes = HTML(string=html, base_url=request.url_root).write_pdf()
    except Exception:
        current_app.logger.exception('PDF generation failed endpoint=%s template=%s rows=%s', request.endpoint, template_name, len(rows))
        flash('PDF could not be generated. Check server logs.', 'danger')
        return redirect(request.referrer or url_for('admin.classes'))

    current_app.logger.info(
        'PDF export generated endpoint=%s template=%s rows=%s bytes=%s',
        request.endpoint,
        template_name,
        len(rows),
        len(pdf_bytes),
    )
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename or "table-export.pdf"}'
    return response


from app.models import (
    School,
    AcademicYear,
    AssessmentSetting,
    FoundationResult,
    FundamentalPupilAttempt,
    FundamentalResponse,
    GapScore,
    Intervention,
    PhonicsScore,
    PhonicsTestColumn,
    Pupil,
    PupilClassHistory,
    ReceptionTrackerEntry,
    SatsColumnResult,
    SatsColumnSetting,
    SatsExamTab,
    SatsResult,
    SatsWritingResult,
    SchoolClass,
    SubjectResult,
    TimesTableScore,
    User,
    WritingResult,
)
from app.services import (
    BOOLEAN_FILTER_CHOICES,
    CLASS_SORT_OPTIONS,
    CORE_SUBJECTS,
    SATS_COLUMN_SUBJECTS,
    RECEPTION_AREAS,
    RECEPTION_STATUS_CHOICES,
    RECEPTION_TRACKING_POINTS,
    FOUNDATION_HALF_TERMS,
    FOUNDATION_JUDGEMENTS,
    FOUNDATION_SUBJECTS,
    FOUNDATION_JUDGEMENT_THEMES,
    SATS_SCORE_TYPES,
    SATS_TRACKER_MODES,
    SUBGROUP_FILTERS,
    TERMS,
    AssessmentValidationError,
    CsvImportError,
    SatsColumnValidationError,
    apply_admin_pupil_filters,
    build_academic_year_options,
    build_admin_pupil_filter_state,
    build_sort_indicator,
    build_table_sort_state,
    build_class_overview_row,
    build_headline_report,
    build_reception_overview,
    build_reception_summary,
    build_reception_tracker_rows,
    build_foundation_summary,
    build_foundation_tracker_rows,
    build_phonics_tracker_rows,
    build_times_tables_tracker_rows,
    build_sats_tracker_rows,
    build_subject_overview_cards,
    build_year6_sats_overview,
    build_next_academic_year,
    build_intervention_filters,
    ensure_academic_year,
    ensure_default_academic_years,
    generate_next_missing_academic_years,
    ensure_default_logins_and_classes,
    export_class_overview_csv,
    export_history_csv,
    export_interventions_csv,
    export_pupil_overview_csv,
    export_reception_tracker_csv,
    export_sats_results_csv,
    export_sats_tracker_csv,
    export_subject_results_csv,
    export_writing_results_csv,
    format_subject_name,
    generate_csv,
    get_class_detail_context,
    get_class_pupil_query,
    get_selected_current_academic_year,
    get_school_working_academic_year,
    get_selected_academic_year,
    get_foundation_half_term,
    get_gender_filter_options,
    get_next_sort_direction,
    get_history_rows,
    get_or_create_assessment_setting,
    get_promotion_mapping_options,
    get_reception_class,
    get_sats_columns,
    get_sats_exam_tabs,
    get_setting_defaults,
    get_tracking_point_key,
    get_tracker_mode,
    get_tracker_mode_label,
    get_current_score_for_intervention,
    import_combined_results,
    is_ks1_year_group,
    is_times_tables_year_group,
    import_reception_tracker,
    import_sats_tracker_results,
    parse_uploaded_csv,
    promote_pupils_to_next_year,
    recalculate_subject_results_for_scope,
    save_sats_column,
    save_phonics_columns,
    save_phonics_scores,
    sort_phonics_tracker_rows,
    save_times_tables_columns,
    save_times_tables_scores,
    sort_times_tables_tracker_rows,
    save_reception_tracker_entries,
    save_foundation_results,
    save_sats_tab,
    set_tracker_mode,
    snapshot_pupil_history,
    sort_class_rows,
    sort_teacher_accounts,
    toggle_sats_column,
    toggle_sats_tab,
    update_assessment_setting,
    ReceptionTrackerValidationError,
    validate_setting_payload,
    add_phonics_column,
    add_times_tables_column,
    ensure_phonics_columns,
    ensure_times_tables_columns,
    FoundationValidationError,
)
from app.utils import admin_required, current_school_id, demo_filter_classes, demo_filter_pupils, is_demo_user, log_audit_event, require_same_school, school_scoped_query
from app.services.pupil_quick_add import create_quick_add_pupil


def _apply_common_report_filters(query):
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    term = request.args.get('term', '').strip()
    class_id = request.args.get('class_id', type=int)
    year_group = request.args.get('year_group', type=int)
    gender = request.args.get('gender', '').strip().lower()
    send = request.args.get('send', 'all').strip().lower()
    pp = request.args.get('pp', 'all').strip().lower()
    laps = request.args.get('laps', 'all').strip().lower()
    service_child = request.args.get('service_child', 'all').strip().lower()

    query = query.join(SchoolClass, Pupil.class_id == SchoolClass.id).filter(Pupil.is_archived.is_(False))
    if class_id:
        query = query.filter(Pupil.class_id == class_id)
    if year_group is not None:
        query = query.filter(SchoolClass.year_group == year_group)
    if gender in {'male', 'female', 'm', 'f'}:
        clause = gender_filter_clause(gender)
        if clause is not None:
            query = query.filter(clause)
    for key, value in [('send', send), ('pupil_premium', pp), ('laps', laps), ('service_child', service_child)]:
        if value in {'yes', 'no'}:
            query = query.filter(getattr(Pupil, key).is_(value == 'yes'))
    return query, {'academic_year':academic_year,'term':term,'class_id':class_id,'year_group':year_group,'gender':gender,'send':send,'pp':pp,'laps':laps,'service_child':service_child}


def _latest_subject_map(pupil_ids):
    if not pupil_ids:
        return {}
    rows = (school_scoped_query(SubjectResult, SubjectResult.query.filter(SubjectResult.pupil_id.in_(pupil_ids)))
        .order_by(SubjectResult.pupil_id, SubjectResult.subject, SubjectResult.term.desc(), SubjectResult.id.desc()).all())
    latest = {}
    for r in rows:
        key=(r.pupil_id,r.subject)
        if key not in latest:
            latest[key]=r.band_label or r.combined_score or '—'
    return latest


def _build_xlsx(headers, rows, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
from app.services.gender import gender_filter_clause, normalize_gender

from . import admin_bp
from .forms import AssessmentSettingForm


def _active_class_query():
    return demo_filter_classes(SchoolClass.query.filter_by(is_active=True))


def _is_active_year6_class(class_id: int | None) -> bool:
    if class_id is None:
        return False
    return demo_filter_classes(SchoolClass.query.filter_by(id=class_id, year_group=6, is_active=True)).first() is not None


def _redirect_non_year6_sats_access():
    flash('Year 6 SATs tracker is only available for Year 6.', 'warning')
    return redirect(url_for('admin.sats'))


def _teacher_options():
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        return []
    teachers = User.query.filter_by(school_id=school_id, role='teacher', is_demo=is_demo_user()).all()
    return sort_teacher_accounts(teachers)




FULL_WORKBOOK_SHEETS = {
    'Instructions': ['Guidance', 'Details'],
    'Pupils': ['pupil_id','Pupil','Class','Year Group','Gender','PP','SEND','LAPS','Service'],
    'Maths': ['pupil_id','Pupil','Class','Year Group','Term','Arithmetic','Reasoning','Notes'],
    'Reading': ['pupil_id','Pupil','Class','Year Group','Term','Paper 1','Paper 2','Notes'],
    'SPaG': ['pupil_id','Pupil','Class','Year Group','Term','Spelling','Grammar','Notes'],
    'Writing': ['pupil_id','Pupil','Class','Year Group','Term','Band','Notes'],
    'Foundation': ['pupil_id','Pupil','Class','Year Group','Subject','Term','Assessment','Band','Notes'],
    'Reception': ['pupil_id','Pupil','Class','Year Group','Term','Area','Statement','Band','Notes'],
    'Phonics': ['pupil_id','Pupil','Class','Year Group','Test Name','Score','Max Score','Date'],
    'Times Tables': ['pupil_id','Pupil','Class','Year Group','Test Name','Score','Max Score','Date'],
    'SATs': ['pupil_id','Pupil','Class','Year Group','Assessment Point','Exam 1','Exam 2','Exam 3','Exam 4','Arithmetic','Reasoning 1','Reasoning 2','Maths Scaled Score','Reading Paper','Reading Scaled','Spelling','Grammar','SPaG Scaled'],
}
TEMPLATE_TERMS = ['Autumn', 'Spring', 'Summer']
FOUNDATION_SUBJECT_KEYS = [subject_key for subject_key, _subject_label in FOUNDATION_SUBJECTS]
SATS_ASSESSMENT_POINTS = ['Autumn 1', 'Autumn 2', 'Spring 1', 'Spring 2']
SATS_FIXED_COLUMNS = {
    'reading_raw': ('reading', 'Reading Raw Score'),
    'reading_scaled': ('reading', 'Reading Scaled Score'),
    'maths_arithmetic_raw': ('maths', 'Arithmetic'),
    'maths_reasoning_raw': ('maths', 'Reasoning'),
    'maths_scaled': ('maths', 'Maths Scaled Score'),
    'spag_grammar_raw': ('spag', 'Grammar'),
    'spag_spelling_raw': ('spag', 'Spelling'),
    'spag_scaled': ('spag', 'SPaG Scaled Score'),
}

def _norm(v):
    return str(v or '').strip()


def _norm_key(v):
    return '_'.join(_norm(v).lower().split())

def _split_name(full_name):
    parts = _norm(full_name).split()
    if not parts:
        return '', ''
    return parts[0], ' '.join(parts[1:]) if len(parts)>1 else ''

def _find_pupil_by_class_name(class_name, pupil_name):
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        return None
    school_class = demo_filter_classes(SchoolClass.query.filter_by(name=_norm(class_name), school_id=school_id)).first()
    if not school_class:
        return None
    first,last=_split_name(pupil_name)
    if not first or not last:
        return None
    return demo_filter_pupils(Pupil.query.filter_by(class_id=school_class.id, first_name=first, last_name=last, school_id=school_id)).first()


def _find_or_create_phonics_column(year_group: int, test_name: str) -> tuple[PhonicsTestColumn, bool]:
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        raise ValueError('Select a school before importing phonics columns.')
    normalized_name = _norm(test_name)
    key = normalized_name.lower()
    existing = (
        PhonicsTestColumn.query
        .filter_by(school_id=school_id, year_group=year_group)
        .all()
    )
    for column in existing:
        if _norm_key(column.name) == key:
            return column, False
    next_order = max((column.display_order or 0 for column in existing), default=0) + 1
    created = PhonicsTestColumn(
        school_id=school_id,
        year_group=year_group,
        name=normalized_name,
        display_order=next_order,
        is_active=True,
    )
    db.session.add(created)
    db.session.flush()
    return created, True


def _find_or_create_sats_column(subject: str, test_name: str) -> tuple[SatsColumnSetting, bool]:
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        raise ValueError('Select a school before importing SATs columns.')
    normalized_name = _norm(test_name)
    key = normalized_name.lower()
    subject_key = _norm_key(subject)
    existing = (
        SatsColumnSetting.query
        .filter_by(school_id=school_id, year_group=6, subject=subject_key)
        .all()
    )
    for column in existing:
        if _norm_key(column.name) == key:
            return column, False

    tabs = (
        SatsExamTab.query
        .filter_by(school_id=school_id, year_group=6, is_active=True)
        .order_by(SatsExamTab.display_order.asc(), SatsExamTab.id.asc())
        .all()
    )
    if not tabs:
        first_tab = SatsExamTab(
            school_id=school_id,
            year_group=6,
            name='Default',
            display_order=1,
            is_active=True,
        )
        db.session.add(first_tab)
        db.session.flush()
        tabs = [first_tab]
    selected_tab = tabs[0]
    next_order = max((column.display_order or 0 for column in existing), default=0) + 1
    created = SatsColumnSetting(
        school_id=school_id,
        year_group=6,
        exam_tab_id=selected_tab.id,
        name=normalized_name,
        subject=subject_key,
        score_type='paper',
        display_order=next_order,
        is_active=True,
    )
    db.session.add(created)
    db.session.flush()
    return created, True


def _normalize_writing_band(value) -> str | None:
    token = _norm(value).lower()
    if token in {'wt', 'wts', 'working towards'}:
        return 'working_towards'
    if token in {'ot', 'on track', 'expected', 'working at'}:
        return 'expected'
    if token in {'exc', 'exs', 'exceeding', 'gds', 'greater depth'}:
        return 'greater_depth'
    return None


def _workbook_effective_school_id() -> int | None:
    """Resolve the school that should own a generated import workbook."""
    if current_user.is_executive_admin:
        return current_school_id()
    return current_user.school_id


def get_import_template_pupils(effective_school_id):
    """Return the single authoritative pupil list for full-school import workbooks."""
    pupils = (
        Pupil.query
        .outerjoin(SchoolClass)
        .filter(Pupil.school_id == effective_school_id)
        .order_by(SchoolClass.year_group, SchoolClass.name, Pupil.name)
        .all()
    )
    return pupils


def _style_template_sheet(ws, header_row: int = 3):
    ws.freeze_panes = f'A{header_row + 1}'
    if ws.max_row >= header_row:
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type='solid', fgColor='E9EFF9')
    for idx, col in enumerate(ws.iter_cols(min_row=header_row, max_row=header_row), start=1):
        header = _norm(col[0].value)
        width = max(14, min(40, len(header) + 4))
        ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].width = width


def _pupil_year_group(pupil: Pupil) -> int | None:
    if pupil.school_class is not None:
        return pupil.school_class.year_group
    return pupil.join_year_group


def _pupil_display_name(pupil: Pupil) -> str:
    return _norm(getattr(pupil, 'name', None)) or pupil.full_name


def _pupil_template_identity(pupil: Pupil) -> tuple[int, str, str, int | None]:
    class_name = pupil.school_class.name if pupil.school_class else ''
    year_group = _pupil_year_group(pupil)
    return pupil.id, _pupil_display_name(pupil), class_name, year_group


def filter_pupils_for_sheet(pupils, sheet_name):
    year_groups_by_sheet = {
        'Reception': {0},
        'Maths': {1, 2, 3, 4, 5},
        'Reading': {1, 2, 3, 4, 5},
        'SPaG': {1, 2, 3, 4, 5},
        'Writing': {1, 2, 3, 4, 5},
        'Foundation': {1, 2, 3, 4, 5},
        'Phonics': {1, 2},
        'Times Tables': {4},
        'SATs': {6},
    }.get(sheet_name)
    if year_groups_by_sheet is None:
        return list(pupils)
    return [pupil for pupil in pupils if _pupil_year_group(pupil) in year_groups_by_sheet]


def _append_template_header(ws, columns):
    ws['A1'] = 'Pupils loaded: 0'
    ws.append([])
    ws.append(columns)
    ws.column_dimensions['A'].hidden = True


def write_pupil_rows(ws, pupils, extra_values_factory=None):
    rows = []
    for pupil in pupils:
        pupil_id, pupil_name, class_name, year_group = _pupil_template_identity(pupil)
        extras = extra_values_factory(pupil) if extra_values_factory else [[]]
        for extra_values in extras:
            rows.append([pupil_id, pupil_name, class_name, year_group, *extra_values])
    for row in rows:
        ws.append(row)
    ws.column_dimensions['A'].hidden = True
    return rows


def _log_full_workbook_export(effective_school_id: int, pupils: list[Pupil], rows_written_by_sheet: dict[str, int]) -> None:
    total_pupils = Pupil.query.count()
    pupils_for_school = Pupil.query.filter(Pupil.school_id == effective_school_id).count()
    current_app.logger.info(
        'Full-school import workbook export: current_user.id=%s current_user.school_id=%s effective_school_id=%s total_pupils_in_db=%s pupils_for_school=%s rows_written_per_sheet=%s',
        current_user.id,
        current_user.school_id,
        effective_school_id,
        total_pupils,
        pupils_for_school,
        rows_written_by_sheet,
    )
    current_app.logger.debug('Full-school import workbook pupil ids: %s', [pupil.id for pupil in pupils])


def _build_full_template_workbook(effective_school_id: int):
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet(title='Instructions')
    _append_template_header(instructions, FULL_WORKBOOK_SHEETS['Instructions'])
    instructions_rows = [
        ('Workbook purpose', 'Use this workbook to import pupil details and assessment data into Class Compass.'),
        ('Do not rename sheets', 'The upload parser uses these exact worksheet names.'),
        ('Do not edit hidden pupil_id columns', 'Hidden pupil IDs are used first to match pupils safely and prevent duplicates.'),
        ('Pupil list source', 'This workbook is generated directly from the selected school database pupils.'),
        ('Academic year', 'Choose the correct academic year on the upload page before importing the completed workbook.'),
        ('Year group routing', 'Reception only appears on Reception; Years 1-5 appear on Maths, Reading, SPaG, Writing, and Foundation; Years 1-2 appear on Phonics; Year 4 appears on Times Tables; Year 6 appears only on SATs.'),
    ]
    for row in instructions_rows:
        instructions.append(row)
    instructions.column_dimensions['A'].width = 26
    instructions.column_dimensions['B'].width = 120

    for sheet_name, columns in FULL_WORKBOOK_SHEETS.items():
        if sheet_name == 'Instructions':
            continue
        _append_template_header(wb.create_sheet(title=sheet_name), columns)

    pupils = get_import_template_pupils(effective_school_id)
    rows_written_by_sheet = {}

    pupils_ws = wb['Pupils']
    for pupil in pupils:
        pupil_id, pupil_name, class_name, year_group = _pupil_template_identity(pupil)
        pupils_ws.append([
            pupil_id,
            pupil_name,
            class_name,
            year_group,
            pupil.gender,
            pupil.pupil_premium,
            pupil.send,
            pupil.laps,
            pupil.service_child,
        ])
    pupils_ws['A1'] = f'Pupils loaded: {len(pupils)}'
    rows_written_by_sheet['Pupils'] = len(pupils)

    extra_values_by_sheet = {
        'Reception': lambda _p: [['', '', '', '', '']],
        'Maths': lambda _p: [[term, '', '', ''] for term in TEMPLATE_TERMS],
        'Reading': lambda _p: [[term, '', '', ''] for term in TEMPLATE_TERMS],
        'SPaG': lambda _p: [[term, '', '', ''] for term in TEMPLATE_TERMS],
        'Writing': lambda _p: [[term, '', ''] for term in TEMPLATE_TERMS],
        'Foundation': lambda _p: [[subject_key, '', '', '', ''] for subject_key in FOUNDATION_SUBJECT_KEYS],
        'Phonics': lambda _p: [['', '', '', '']],
        'Times Tables': lambda _p: [['', '', '', '']],
        'SATs': lambda _p: [[assessment_point, '', '', '', '', '', '', '', '', '', '', '', '', ''] for assessment_point in SATS_ASSESSMENT_POINTS],
    }
    for sheet_name, extra_values_factory in extra_values_by_sheet.items():
        sheet_pupils = filter_pupils_for_sheet(pupils, sheet_name)
        rows = write_pupil_rows(wb[sheet_name], sheet_pupils, extra_values_factory=extra_values_factory)
        wb[sheet_name]['A1'] = f'Pupils loaded: {len(sheet_pupils)}'
        rows_written_by_sheet[sheet_name] = len(rows)

    for sheet_name in FULL_WORKBOOK_SHEETS:
        _style_template_sheet(wb[sheet_name], header_row=3)

    _log_full_workbook_export(effective_school_id, pupils, rows_written_by_sheet)
    return wb


CLASS_DETAIL_SUBJECT_SORT_COLUMNS = {'name', 'paper_1_score', 'paper_2_score', 'combined_score', 'combined_percent', 'band_label', 'assessment_year_group', 'progress_delta'}
CLASS_DETAIL_WRITING_SORT_COLUMNS = {'name', 'band_label', 'notes'}
PUPIL_STATUS_FILTER_CHOICES = (
    ('active', 'Active pupils only'),
    ('all', 'Include archived pupils'),
    ('archived', 'Archived pupils only'),
)
PUPIL_LINKED_MODELS = (
    ('subject results', SubjectResult),
    ('writing results', WritingResult),
    ('GAP scores', GapScore),
    ('interventions', Intervention),
    ('SATs results', SatsResult),
    ('SATs writing results', SatsWritingResult),
    ('SATs column results', SatsColumnResult),
    ('phonics scores', PhonicsScore),
    ('times tables scores', TimesTableScore),
    ('foundation judgements', FoundationResult),
    ('reception tracker entries', ReceptionTrackerEntry),
    ('class history records', PupilClassHistory),
)


def _linked_pupil_record_counts(pupil_id: int) -> dict[str, int]:
    counts = {
        label: model.query.filter_by(pupil_id=pupil_id).count()
        for label, model in PUPIL_LINKED_MODELS
    }
    counts['Maths Fundamentals attempts'] = FundamentalPupilAttempt.query.filter_by(pupil_id=pupil_id).count()
    return counts


def _delete_pupil_linked_data(pupil: Pupil) -> None:
    attempt_ids = FundamentalPupilAttempt.query.with_entities(FundamentalPupilAttempt.id).filter_by(pupil_id=pupil.id)
    FundamentalResponse.query.filter(FundamentalResponse.attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
    FundamentalPupilAttempt.query.filter_by(pupil_id=pupil.id).delete(synchronize_session=False)
    for _, model in PUPIL_LINKED_MODELS:
        model.query.filter_by(pupil_id=pupil.id, school_id=pupil.school_id).delete(synchronize_session=False)


def _linked_record_summary(linked_counts: dict[str, int]) -> str:
    populated = [f'{label}: {count}' for label, count in linked_counts.items() if count]
    return ', '.join(populated)


def _class_linked_data_counts(school_class: SchoolClass) -> dict[str, int]:
    pupil_ids_subquery = Pupil.query.with_entities(Pupil.id).filter_by(class_id=school_class.id).subquery()
    return {
        'pupils': Pupil.query.filter_by(class_id=school_class.id).count(),
        'subject_results': SubjectResult.query.filter(SubjectResult.pupil_id.in_(pupil_ids_subquery)).count(),
        'writing_results': WritingResult.query.filter(WritingResult.pupil_id.in_(pupil_ids_subquery)).count(),
        'interventions': Intervention.query.filter(Intervention.pupil_id.in_(pupil_ids_subquery)).count(),
        'sats_results': SatsResult.query.filter(SatsResult.pupil_id.in_(pupil_ids_subquery)).count(),
        'sats_writing_results': SatsWritingResult.query.filter(SatsWritingResult.pupil_id.in_(pupil_ids_subquery)).count(),
        'sats_column_results': SatsColumnResult.query.filter(SatsColumnResult.pupil_id.in_(pupil_ids_subquery)).count(),
        'phonics_scores': PhonicsScore.query.filter(PhonicsScore.pupil_id.in_(pupil_ids_subquery)).count(),
        'times_table_scores': TimesTableScore.query.filter(TimesTableScore.pupil_id.in_(pupil_ids_subquery)).count(),
        'reception_tracker_entries': ReceptionTrackerEntry.query.filter(ReceptionTrackerEntry.pupil_id.in_(pupil_ids_subquery)).count(),
        'pupil_class_history': PupilClassHistory.query.filter(PupilClassHistory.pupil_id.in_(pupil_ids_subquery)).count(),
        'foundation_results': FoundationResult.query.filter(FoundationResult.pupil_id.in_(pupil_ids_subquery)).count(),
    }


def _pupil_action_redirect():
    next_url = request.form.get('next', '').strip()
    return redirect(next_url or url_for('admin.pupils'))


def _school_scope_filter(model):
    school_id = current_school_id()
    if school_id is not None and hasattr(model, 'school_id'):
        return model.school_id == school_id
    return True


def _table_header_state(sort_state: dict, allowed_columns: set[str]) -> dict:
    return {
        column: {
            'indicator': build_sort_indicator(column, sort_state),
            'next_direction': get_next_sort_direction(column, sort_state),
            'active': sort_state['column'] == column,
        }
        for column in allowed_columns
    }


@admin_bp.route('/classes', methods=['GET', 'POST'])
@login_required
@admin_required
def classes():
    effective_school_id = _selected_school_id_for_admin_actions()

    if request.method == 'POST':
        action = request.form.get('action', 'create_class')
        if effective_school_id is None:
            flash('Select a school before managing classes.', 'warning')
            return redirect(url_for('admin.classes'))
        if is_demo_user() and action == 'archive_class':
            flash('This action is disabled in Demo Mode.', 'warning')
            return redirect(url_for('admin.classes'))
        try:
            if action == 'create_class':
                name = request.form.get('name', '').strip()
                year_group = int(request.form.get('year_group', '0'))
                teacher_id_raw = request.form.get('teacher_id', '').strip()
                if not name:
                    raise ValueError('Class name is required.')
                existing = demo_filter_classes(SchoolClass.query).filter_by(school_id=effective_school_id, name=name).first()
                if existing:
                    raise ValueError('A class with that name already exists in your school.')
                school_class = SchoolClass(name=name, year_group=year_group, school_id=effective_school_id)
                school_class.teacher_id = int(teacher_id_raw) if teacher_id_raw else None
                school_class.is_active = True
                school_class.is_demo = current_user.is_demo
                db.session.add(school_class)
                flash(f'Created class {name}.', 'success')
            elif action == 'update_class':
                school_class = SchoolClass.query.get_or_404(int(request.form.get('class_id', '0')))
                require_same_school(school_class)
                new_name = request.form.get(f'name_{school_class.id}', '').strip()
                new_year_group = int(request.form.get(f'year_group_{school_class.id}', school_class.year_group))
                teacher_id_raw = request.form.get(f'teacher_id_{school_class.id}', '').strip()
                if not new_name:
                    raise ValueError('Class name is required.')
                existing = demo_filter_classes(SchoolClass.query).filter(
                    SchoolClass.school_id == school_class.school_id,
                    SchoolClass.name == new_name,
                    SchoolClass.id != school_class.id,
                ).first()
                if existing:
                    raise ValueError('A class with that name already exists in this school.')
                school_class.name = new_name
                school_class.year_group = new_year_group
                school_class.teacher_id = int(teacher_id_raw) if teacher_id_raw else None
                school_class.is_active = request.form.get(f'is_active_{school_class.id}') == 'on'
                db.session.add(school_class)
                flash(f'Updated class {school_class.name}.', 'success')
            elif action == 'archive_class':
                school_class = demo_filter_classes(SchoolClass.query).filter_by(
                    id=int(request.form.get('class_id', '0')),
                    school_id=effective_school_id,
                ).first_or_404()
                school_class.is_active = False
                if hasattr(school_class, 'is_archived'):
                    school_class.is_archived = True
                db.session.add(school_class)
                current_app.logger.info('Class action: archived class_id=%s school_id=%s', school_class.id, school_class.school_id)
                flash(f'Archived class {school_class.name}.', 'success')
            elif action == 'restore_class':
                school_class = demo_filter_classes(SchoolClass.query).filter_by(
                    id=int(request.form.get('class_id', '0')),
                    school_id=effective_school_id,
                ).first_or_404()
                school_class.is_active = True
                if hasattr(school_class, 'is_archived'):
                    school_class.is_archived = False
                db.session.add(school_class)
                flash(f'Restored class {school_class.name}.', 'success')
            elif action == 'delete_class':
                school_class = demo_filter_classes(SchoolClass.query).filter_by(
                    id=int(request.form.get('class_id', '0')),
                    school_id=effective_school_id,
                ).first_or_404()
                linked_counts = _class_linked_data_counts(school_class)
                if any(linked_counts.values()):
                    current_app.logger.info('Class action: delete_blocked class_id=%s school_id=%s linked=%s', school_class.id, school_class.school_id, linked_counts)
                    raise ValueError('This class has pupils or linked data. Archive it instead.')
                class_name = school_class.name
                db.session.delete(school_class)
                current_app.logger.info('Class action: hard_deleted class_id=%s school_id=%s', school_class.id, school_class.school_id)
                flash(f'Deleted empty class {class_name}.', 'success')
            db.session.commit()
            return redirect(url_for('admin.classes'))
        except ValueError as exc:
            db.session.rollback()
            flash(f'Class changes could not be saved: {exc}', 'danger')

    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    filter_year_group = request.args.get('year_group', '').strip()
    filter_teacher = request.args.get('teacher_id', '').strip()
    filter_class = request.args.get('class_id', '').strip()
    show_archived = request.args.get('show_archived', '0').strip() == '1'
    send_filter = (request.args.get('send', 'all') or 'all').strip().lower()
    sort = request.args.get('sort', 'year_group')

    query = demo_filter_classes(SchoolClass.query).filter(SchoolClass.is_active.is_(True))
    if effective_school_id is not None:
        query = query.filter(SchoolClass.school_id == effective_school_id)
    elif current_user.is_executive_admin:
        query = query.filter(False)
    if filter_year_group:
        query = query.filter(SchoolClass.year_group == int(filter_year_group))
    if filter_teacher:
        query = query.filter(SchoolClass.teacher_id == int(filter_teacher))
    if filter_class:
        query = query.filter(SchoolClass.id == int(filter_class))

    classes = query.order_by(SchoolClass.year_group, SchoolClass.name).all()
    rows = [build_class_overview_row(school_class, academic_year, filters={'send': send_filter}) for school_class in classes]
    rows = sort_class_rows(rows, sort)
    return render_template(
        'admin/classes.html',
        classes=rows,
        academic_year=academic_year,
        filter_year_group=filter_year_group,
        filter_teacher=filter_teacher,
        filter_class=filter_class,
        show_archived=show_archived,
        sort=sort,
        send_filter=send_filter,
        sort_options=CLASS_SORT_OPTIONS,
        teacher_options=_teacher_options(),
        class_options=demo_filter_classes(SchoolClass.query).filter(SchoolClass.is_active.is_(True), SchoolClass.school_id == effective_school_id).order_by(SchoolClass.year_group, SchoolClass.name).all() if effective_school_id is not None else [],
        archived_classes=demo_filter_classes(SchoolClass.query).filter(SchoolClass.is_active.is_(False), SchoolClass.school_id == effective_school_id).order_by(SchoolClass.year_group, SchoolClass.name).all() if show_archived and effective_school_id is not None else [],
    )


@admin_bp.route('/classes/<int:class_id>')
@login_required
@admin_required
def class_detail(class_id: int):
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == class_id).first_or_404()
    require_same_school(school_class)
    pupil_filters = build_admin_pupil_filter_state(request.args)
    selected_subject = request.args.get('subject', 'maths').strip() or 'maths'
    selected_term = request.args.get('term', '').strip() or None
    allowed_columns = CLASS_DETAIL_WRITING_SORT_COLUMNS if selected_subject == 'writing' else CLASS_DETAIL_SUBJECT_SORT_COLUMNS
    sort_state = build_table_sort_state(request.args, allowed_columns=allowed_columns, default_column='name')
    context = get_class_detail_context(
        school_class,
        academic_year,
        subject=selected_subject,
        term=selected_term,
        filters=pupil_filters,
        sort_column=sort_state['column'],
        sort_direction=sort_state['direction'],
    )
    if context['selected_subject'] == 'writing':
        header_state = _table_header_state(sort_state, CLASS_DETAIL_WRITING_SORT_COLUMNS)
    elif context['selected_subject'] in {'maths', 'reading', 'spag'}:
        header_state = _table_header_state(sort_state, CLASS_DETAIL_SUBJECT_SORT_COLUMNS)
    else:
        header_state = {}

    export_mode = request.args.get('export', '').strip().lower() == 'csv'
    print_mode = request.args.get('print', '0') == '1'
    pdf_mode = _pdf_requested()
    anon_mode = request.args.get('anon', '0') == '1'
    if export_mode or print_mode or pdf_mode:
        headers = ['Pupil', 'Class', 'Gender', 'PP', 'SEND', 'LAPS', 'Service']
        rows = []
        for idx, row in enumerate(context['pupil_rows'], start=1):
            pupil_name = f'Pupil {idx}' if anon_mode else row['name']
            rows.append([pupil_name, school_class.name, normalize_gender(row.get('gender')) or '', 'Yes' if row.get('pupil_premium') else 'No', 'Yes' if row.get('send') else 'No', 'Yes' if row.get('laps') else 'No', 'Yes' if row.get('service_child') else 'No'])
        if context['selected_subject'] in {'maths', 'reading', 'spag'}:
            headers += ['Paper 1', 'Paper 2', 'Combined', 'Band']
            for i, row in enumerate(context['pupil_rows']):
                rows[i] += [row.get('paper_1_score'), row.get('paper_2_score'), row.get('combined_score'), row.get('band_label') or '']
        elif context['selected_subject'] == 'writing':
            headers += ['Writing judgement', 'Notes']
            for i, row in enumerate(context['pupil_rows']):
                rows[i] += [row.get('band_label') or '', row.get('notes') or '']
        if export_mode:
            out = io.StringIO(); w = csv.writer(out); w.writerow(headers); w.writerows(rows)
            return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename=class_{context["selected_subject"]}_export.csv'})
        title = f'{context["subject_label"]} class report'
        subtitle = 'Anonymised' if anon_mode else 'Named'
        if pdf_mode:
            return _render_table_pdf(title, headers, rows, pupil_filters, anon_mode, f'class_{context["selected_subject"]}_report.pdf', subtitle)
        return render_template('admin/report_table.html', title=title, subtitle=subtitle, headers=headers, rows=rows, filters=pupil_filters, anonymised=anon_mode)

    return render_template(
        'admin/class_detail.html',
        academic_year=academic_year,
        boolean_filter_choices=BOOLEAN_FILTER_CHOICES,
        gender_options=get_gender_filter_options(
            class_id=school_class.id,
            include_inactive=pupil_filters.get('pupil_status') != 'active',
        ),
        pupil_status_filter_choices=PUPIL_STATUS_FILTER_CHOICES,
        sort_state=sort_state,
        header_state=header_state,
        **context,
    )


@admin_bp.route('/classes/<int:class_id>/sats')
@login_required
@admin_required
def class_sats(class_id: int):
    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == class_id).first_or_404()
    require_same_school(school_class)
    return redirect(url_for('dashboards.sats_simple', class_id=class_id, school_id=school_class.school_id))

# legacy disabled
def _legacy_class_sats_disabled(class_id: int):
    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == class_id).first_or_404()
    require_same_school(school_class)
    if school_class.year_group != 6 or not school_class.is_active:
        return _redirect_non_year6_sats_access()
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    selected_tab_id_raw = request.args.get('exam_tab_id', '').strip()
    pupils = get_class_pupil_query(school_class, academic_year).filter(Pupil.is_active.is_(True), Pupil.school_id == school_class.school_id).order_by(Pupil.last_name, Pupil.first_name).all()
    columns, rows, overview = build_sats_tracker_rows(pupils, academic_year, 6, exam_tab_id=int(selected_tab_id_raw) if selected_tab_id_raw else None, active_only=True)
    selected_tab = overview.pop('_selected_tab', None)
    tabs = overview.pop('_tabs', get_sats_exam_tabs(6, include_inactive=True))
    return render_template(
        'admin/sats.html',
        academic_year=academic_year,
        tracker_mode=get_tracker_mode(6),
        tracker_mode_label=get_tracker_mode_label(6),
        tracker_mode_options=SATS_TRACKER_MODES,
        class_options=demo_filter_classes(SchoolClass.query.filter_by(year_group=6, is_active=True)).order_by(SchoolClass.name).all(),
        selected_class_id=school_class.id,
        columns=columns,
        all_columns=get_sats_columns(6, exam_tab_id=selected_tab.id if selected_tab else None, active_only=False),
        tabs=tabs,
        selected_tab=selected_tab,
        rows=rows,
        overview=overview,
        class_summaries=[{'class': school_class, 'rows': rows, 'subject_totals': overview}],
        sats_subject_choices=SATS_COLUMN_SUBJECTS,
        sats_score_type_choices=SATS_SCORE_TYPES,
    )


@admin_bp.route('/classes/<int:class_id>/phonics', methods=['GET', 'POST'])
@login_required
@admin_required
def class_phonics(class_id: int):
    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == class_id).first_or_404()
    require_same_school(school_class)
    academic_year = request.values.get('academic_year', get_selected_current_academic_year())
    filters = build_admin_pupil_filter_state(request.values)

    if not is_ks1_year_group(school_class.year_group):
        flash('The phonics tracker is only available for Year 1 and Year 2 classes.', 'warning')
        return redirect(url_for('admin.class_detail', class_id=class_id, academic_year=academic_year))

    pupils = apply_admin_pupil_filters(get_class_pupil_query(school_class, academic_year).filter(Pupil.is_active.is_(True)), filters).order_by(Pupil.last_name, Pupil.first_name).all()
    columns = ensure_phonics_columns(school_class.year_group, school_class.school_id)
    active_columns = [column for column in columns if column.is_active]
    sortable_columns = {'name', *(f'column_{column.id}' for column in active_columns)}
    sort_state = build_table_sort_state(request.values, allowed_columns=sortable_columns, default_column='name')
    header_state = {
        column: {
            'indicator': build_sort_indicator(column, sort_state),
            'next_direction': get_next_sort_direction(column, sort_state),
            'active': sort_state['column'] == column,
        }
        for column in sortable_columns
    }

    if request.method == 'POST':
        action = request.form.get('action', 'save_scores')
        try:
            if action == 'save_columns':
                columns = save_phonics_columns(school_class.year_group, school_class.school_id, request.form)
                flash('Phonics test columns updated.', 'success')
            elif action == 'add_column':
                column = add_phonics_column(school_class.year_group, school_class.school_id, request.form)
                flash(f'Added phonics column {column.name}.', 'success')
            else:
                save_phonics_scores(pupils, columns, academic_year, school_class.school_id, request.form)
                flash('Phonics scores saved.', 'success')
            db.session.commit()
            return redirect(url_for('admin.class_phonics', class_id=class_id, academic_year=academic_year, pupil_status=filters['pupil_status'], gender=filters['gender'], pupil_premium=filters['pupil_premium'], laps=filters['laps'], service_child=filters['service_child'], send=filters['send'], search=filters['search'], sort=sort_state['column'], direction=sort_state['direction']))
        except ValueError as exc:
            db.session.rollback()
            flash(f'Phonics changes could not be saved: {exc}', 'danger')
            columns = ensure_phonics_columns(school_class.year_group, school_class.school_id)

    rows = build_phonics_tracker_rows(pupils, columns, academic_year, school_class.school_id)
    rows = sort_phonics_tracker_rows(rows, sort_state['column'], sort_state['direction'])
    return render_template(
        'admin/class_phonics.html',
        school_class=school_class,
        columns=columns,
        rows=rows,
        pupils=pupils,
        academic_year=academic_year,
        academic_year_options=build_academic_year_options(academic_year),
        filters=filters,
        boolean_filter_choices=BOOLEAN_FILTER_CHOICES,
        gender_options=get_gender_filter_options(
            class_id=school_class.id,
            include_inactive=filters.get('pupil_status') != 'active',
        ),
        pupil_status_filter_choices=PUPIL_STATUS_FILTER_CHOICES,
        sort_state=sort_state,
        header_state=header_state,
    )


@admin_bp.route('/classes/<int:class_id>/times_tables', methods=['GET', 'POST'])
@login_required
@admin_required
def class_times_tables(class_id: int):
    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == class_id).first_or_404()
    require_same_school(school_class)
    academic_year = request.values.get('academic_year', get_selected_current_academic_year())
    filters = build_admin_pupil_filter_state(request.values)

    if not is_times_tables_year_group(school_class.year_group):
        flash('The times tables tracker is only available for Year 4 classes.', 'warning')
        return redirect(url_for('admin.class_detail', class_id=class_id, academic_year=academic_year))

    pupils = apply_admin_pupil_filters(get_class_pupil_query(school_class, academic_year).filter(Pupil.is_active.is_(True)), filters).order_by(Pupil.last_name, Pupil.first_name).all()
    columns = ensure_times_tables_columns(school_class.year_group)
    active_columns = [column for column in columns if column.is_active]
    sortable_columns = {'name', *(f'column_{column.id}' for column in active_columns)}
    sort_state = build_table_sort_state(request.values, allowed_columns=sortable_columns, default_column='name')
    header_state = {
        column: {
            'indicator': build_sort_indicator(column, sort_state),
            'next_direction': get_next_sort_direction(column, sort_state),
            'active': sort_state['column'] == column,
        }
        for column in sortable_columns
    }

    if request.method == 'POST':
        action = request.form.get('action', 'save_scores')
        try:
            if action == 'save_columns':
                columns = save_times_tables_columns(school_class.year_group, request.form)
                flash('Times tables test columns updated.', 'success')
            elif action == 'add_column':
                column = add_times_tables_column(school_class.year_group, request.form)
                flash(f'Added times tables column {column.name}.', 'success')
            else:
                save_times_tables_scores(pupils, columns, academic_year, request.form)
                flash('Times tables scores saved.', 'success')
            db.session.commit()
            return redirect(url_for('admin.class_times_tables', class_id=class_id, academic_year=academic_year, pupil_status=filters['pupil_status'], gender=filters['gender'], pupil_premium=filters['pupil_premium'], laps=filters['laps'], service_child=filters['service_child'], send=filters['send'], search=filters['search'], sort=sort_state['column'], direction=sort_state['direction']))
        except ValueError as exc:
            db.session.rollback()
            flash(f'Times tables changes could not be saved: {exc}', 'danger')
            columns = ensure_times_tables_columns(school_class.year_group)

    rows = build_times_tables_tracker_rows(pupils, columns, academic_year)
    rows = sort_times_tables_tracker_rows(rows, sort_state['column'], sort_state['direction'])
    return render_template(
        'admin/class_times_tables.html',
        school_class=school_class,
        columns=columns,
        rows=rows,
        pupils=pupils,
        academic_year=academic_year,
        academic_year_options=build_academic_year_options(academic_year),
        filters=filters,
        boolean_filter_choices=BOOLEAN_FILTER_CHOICES,
        gender_options=get_gender_filter_options(
            class_id=school_class.id,
            include_inactive=filters.get('pupil_status') != 'active',
        ),
        pupil_status_filter_choices=PUPIL_STATUS_FILTER_CHOICES,
        sort_state=sort_state,
        header_state=header_state,
    )


@admin_bp.route('/foundation', methods=['GET', 'POST'])
@login_required
@admin_required
def foundation_tracker():
    class_options = demo_filter_classes(SchoolClass.query.filter_by(is_active=True)).order_by(SchoolClass.year_group, SchoolClass.name).all()
    if not class_options:
        flash('No active classes are available yet.', 'warning')
        return redirect(url_for('admin.classes'))

    selected_class_id_raw = (request.values.get('class_id') or '').strip()
    school_class = next((item for item in class_options if str(item.id) == selected_class_id_raw), class_options[0])
    academic_year = request.values.get('academic_year', get_selected_current_academic_year())
    half_term = get_foundation_half_term(request.values.get('half_term'))
    filters = build_admin_pupil_filter_state(request.values)
    pupils_query = get_class_pupil_query(school_class, academic_year)
    pupils = apply_admin_pupil_filters(pupils_query, filters).order_by(Pupil.last_name, Pupil.first_name).all()

    if request.method == 'POST':
        half_term = get_foundation_half_term(request.form.get('half_term'))
        try:
            save_foundation_results(pupils, academic_year, half_term, request.form, user_id=current_user.id)
            db.session.commit()
            flash('Foundation judgements saved.', 'success')
            return redirect(
                url_for(
                    'admin.foundation_tracker',
                    class_id=school_class.id,
                    academic_year=academic_year,
                    half_term=half_term,
                    pupil_status=filters['pupil_status'],
                    gender=filters['gender'],
                    pupil_premium=filters['pupil_premium'],
                    laps=filters['laps'],
                    service_child=filters['service_child'],
                    search=filters['search'],
                )
            )
        except FoundationValidationError as exc:
            db.session.rollback()
            flash(f'Foundation changes could not be saved: {exc}', 'danger')

    rows = build_foundation_tracker_rows(pupils, academic_year, half_term)
    summary = build_foundation_summary(rows)
    return render_template(
        'admin/foundation_tracker.html',
        school_class=school_class,
        class_options=class_options,
        rows=rows,
        summary=summary,
        academic_year=academic_year,
        academic_year_options=build_academic_year_options(academic_year),
        half_terms=FOUNDATION_HALF_TERMS,
        selected_half_term=half_term,
        subjects=FOUNDATION_SUBJECTS,
        judgement_choices=FOUNDATION_JUDGEMENTS,
        judgement_themes=FOUNDATION_JUDGEMENT_THEMES,
        filters=filters,
        boolean_filter_choices=BOOLEAN_FILTER_CHOICES,
        gender_options=get_gender_filter_options(
            class_id=school_class.id,
            include_inactive=filters.get('pupil_status') != 'active',
        ),
        pupil_status_filter_choices=PUPIL_STATUS_FILTER_CHOICES,
    )


@admin_bp.route('/reception', methods=['GET', 'POST'])
@login_required
@admin_required
def reception_tracker():
    school_class = get_reception_class()
    if not school_class:
        flash('Reception class has not been created yet. Open Users and run Sync defaults or create Reception class.', 'warning')
        return redirect(url_for('admin.classes'))

    academic_year = request.values.get('academic_year', get_selected_current_academic_year())
    tracking_point = get_tracking_point_key(request.values.get('tracking_point'))
    view = (request.values.get('view', 'tracker') or 'tracker').strip().lower()
    if view not in {'tracker', 'overview'}:
        view = 'tracker'
    pupils = get_class_pupil_query(school_class, academic_year).filter(Pupil.is_active.is_(True)).order_by(Pupil.last_name, Pupil.first_name).all()

    if request.method == 'POST':
        tracking_point = get_tracking_point_key(request.form.get('tracking_point'))
        try:
            save_reception_tracker_entries(pupils, academic_year, tracking_point, request.form)
            db.session.commit()
            flash(f'Reception tracker saved for {dict(RECEPTION_TRACKING_POINTS)[tracking_point]}.', 'success')
            return redirect(url_for('admin.reception_tracker', academic_year=academic_year, tracking_point=tracking_point, view=view))
        except ReceptionTrackerValidationError as exc:
            db.session.rollback()
            flash(f'Reception tracker could not be saved: {exc}', 'danger')

    rows = build_reception_tracker_rows(pupils, academic_year, tracking_point)
    summary = build_reception_summary(rows)
    overview = build_reception_overview(rows)
    return render_template(
        'admin/reception_tracker.html',
        school_class=school_class,
        academic_year=academic_year,
        academic_year_options=build_academic_year_options(academic_year),
        tracking_points=RECEPTION_TRACKING_POINTS,
        selected_tracking_point=tracking_point,
        areas=RECEPTION_AREAS,
        status_choices=RECEPTION_STATUS_CHOICES,
        rows=rows,
        summary=summary,
        overview=overview,
        selected_view=view,
    )


@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def users():
    effective_school_id = _selected_school_id_for_admin_actions()

    if request.method == 'POST':
        action = request.form.get('action', 'create')
        if is_demo_user() and action == 'delete':
            flash('This action is disabled in Demo Mode.', 'warning')
            return redirect(url_for('admin.users'))
        try:
            if effective_school_id is None:
                raise ValueError('Select a school before managing users.')
            if action == 'create':
                username = request.form.get('username', '').strip()
                password = request.form.get('password', '').strip()
                class_id_raw = request.form.get('class_id', '').strip()
                if not username or not password:
                    raise ValueError('Username and password are required.')
                if User.query.filter(
                    User.school_id == effective_school_id,
                    User.username.ilike(username),
                ).first():
                    raise ValueError('That username already exists.')
                user = User(
                    username=username,
                    role='teacher',
                    legacy_is_admin=False,
                    is_active=True,
                    is_demo=is_demo_user(),
                    school_id=effective_school_id,
                )
                user.set_password(password)
                db.session.add(user)
                db.session.flush()
                if class_id_raw:
                    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == int(class_id_raw), SchoolClass.school_id == effective_school_id).first_or_404()
                    require_same_school(school_class)
                    school_class.teacher_id = user.id
                    db.session.add(school_class)
                flash(f'Created teacher user {username}.', 'success')
            elif action == 'update':
                user = User.query.get_or_404(int(request.form.get('user_id', '0')))
                require_same_school(user)
                username = request.form.get(f'username_{user.id}', '').strip()
                class_id_raw = request.form.get(f'class_id_{user.id}', '').strip()
                if username and username != user.username:
                    if User.query.filter(
                        User.school_id == effective_school_id,
                        User.username.ilike(username),
                        User.id != user.id,
                    ).first():
                        raise ValueError('That username is already in use.')
                    user.username = username
                user.is_active = request.form.get(f'is_active_{user.id}') == 'on'
                for school_class in user.classes.all():
                    if not class_id_raw or school_class.id != int(class_id_raw):
                        school_class.teacher_id = None
                        db.session.add(school_class)
                if class_id_raw:
                    school_class = demo_filter_classes(SchoolClass.query).filter(SchoolClass.id == int(class_id_raw), SchoolClass.school_id == effective_school_id).first_or_404()
                    require_same_school(school_class)
                    school_class.teacher_id = user.id
                    db.session.add(school_class)
                db.session.add(user)
                flash(f'Updated {user.username}.', 'success')
            elif action == 'delete':
                user = User.query.get_or_404(int(request.form.get('user_id', '0')))
                require_same_school(user)
                if user.id == current_user.id:
                    raise ValueError('You cannot delete your own account while logged in.')
                if user.can_manage_school and school_scoped_query(User, User.query.filter(User.role.in_(['school_admin','admin']), User.is_active.is_(True))).count() <= 1:
                    raise ValueError('You cannot delete the last remaining active admin user.')
                for school_class in user.classes.all():
                    school_class.teacher_id = None
                    db.session.add(school_class)
                db.session.delete(user)
                flash(f'Deleted user {user.username}.', 'success')
            elif action == 'sync_defaults':
                if not current_app.config.get('ALLOW_DEV_BOOTSTRAP', False):
                    raise ValueError('Default account sync is disabled in production.')
                ensure_default_logins_and_classes()
                flash('Default development accounts and Year 1–6 class links were refreshed.', 'success')
            db.session.commit()
            return redirect(url_for('admin.users'))
        except ValueError as exc:
            db.session.rollback()
            flash(f'User changes could not be saved: {exc}', 'danger')

    users_query = User.query
    if effective_school_id is not None:
        users_query = users_query.filter(User.school_id == effective_school_id)
    elif current_user.role == 'executive_admin':
        users_query = users_query.filter(False)
    teachers = sort_teacher_accounts(users_query.order_by(User.role.desc(), User.username).all())
    classes = demo_filter_classes(SchoolClass.query).filter(SchoolClass.school_id == effective_school_id).order_by(SchoolClass.year_group, SchoolClass.name).all() if effective_school_id is not None else []
    return render_template(
        'admin/users.html',
        teachers=teachers,
        classes=classes,
        active_admin_count=users_query.filter(User.role.in_(['school_admin', 'admin']), User.is_active.is_(True)).count(),
        allow_dev_bootstrap=current_app.config.get('ALLOW_DEV_BOOTSTRAP', False),
    )


@admin_bp.route('/users/<int:user_id>/reset-password', methods=['GET', 'POST'])
@login_required
@admin_required
def reset_user_password(user_id: int):
    user = User.query.get_or_404(user_id)
    require_same_school(user)
    if user.is_executive_admin:
        flash('Admin passwords cannot be reset from this page.', 'warning')
        return redirect(url_for('admin.users'))

    if request.method == 'POST':
        password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        require_change = request.form.get('require_password_change') == 'on'
        if not password:
            flash('A new password is required.', 'danger')
        elif len(password) < 8:
            flash('New password must be at least 8 characters long.', 'danger')
        elif password != confirm_password:
            flash('Passwords did not match.', 'danger')
        else:
            user.set_password(password)
            user.require_password_change = require_change
            db.session.add(user)
            db.session.commit()
            flash(f'Password reset for {user.username}.', 'success')
            return redirect(url_for('admin.users'))

    return render_template('admin/reset_password.html', user=user)




@admin_bp.route('/pupils/new', methods=['GET', 'POST'], endpoint='new_pupil')
@admin_bp.route('/pupils/new', methods=['GET', 'POST'], endpoint='admin_pupil_new')
@login_required
@admin_required
def admin_pupil_new():
    school_id = _selected_school_id_for_admin_actions()
    if current_user.is_executive_admin and school_id is None:
        flash('Select a school before adding pupils.', 'warning')
        return redirect(url_for('admin.pupils'))

    class_options = demo_filter_classes(
        SchoolClass.query.filter_by(school_id=school_id, is_active=True)
    ).order_by(SchoolClass.year_group, SchoolClass.name).all() if school_id is not None else []

    if request.method == 'POST':
        school_class = _resolve_admin_quick_add_class(request.form.get('class_id', '0'))
        pupil, error = create_quick_add_pupil(
            school_class=school_class,
            first_name=request.form.get('first_name', ''),
            last_name=request.form.get('last_name', ''),
            gender=request.form.get('gender', ''),
            pupil_premium=request.form.get('pupil_premium') == 'on',
            laps=request.form.get('laps') == 'on',
            service_child=request.form.get('service_child') == 'on',
            send=request.form.get('send') == 'on',
            join_year_group_raw=request.form.get('join_year_group', ''),
            join_date_raw=request.form.get('join_date', ''),
        )
        if error:
            flash(error, 'danger')
            return render_template('admin/pupil_form.html', class_options=class_options)

        pupil.general_notes = request.form.get('general_notes', '').strip() or None
        db.session.add(pupil)
        db.session.commit()
        flash('Pupil added.', 'success')
        return redirect(url_for('admin.pupils'))

    return render_template('admin/pupil_form.html', class_options=class_options)

@admin_bp.route('/pupils', methods=['GET', 'POST'])
@login_required
@admin_required
def pupils():
    if request.method == 'POST':
        return _handle_admin_quick_add_form_redirect()

    pupil_filters = build_admin_pupil_filter_state(request.args)
    class_id_raw = request.args.get('class_id', '').strip()
    send_filter = (request.args.get('send', 'all') or 'all').strip().lower()

    query = demo_filter_pupils(apply_admin_pupil_filters(Pupil.query, pupil_filters))
    if class_id_raw:
        query = query.filter(Pupil.class_id == int(class_id_raw))
    if send_filter == 'yes':
        query = query.filter(Pupil.send.is_(True))
    elif send_filter == 'no':
        query = query.filter(or_(Pupil.send.is_(False), Pupil.send.is_(None)))
    pupils = query.order_by(Pupil.last_name, Pupil.first_name).all()
    if _pdf_requested() or request.args.get('print', '0') == '1':
        anon_mode = request.args.get('anon', '0') == '1'
        headers = ['Pupil', 'Class', 'Gender', 'PP', 'SEND', 'LAPS', 'Service']
        rows = [[f'Pupil {idx}' if anon_mode else pupil.full_name, pupil.school_class.name if pupil.school_class else '', normalize_gender(pupil.gender) or '', 'Yes' if pupil.pupil_premium else 'No', 'Yes' if pupil.send else 'No', 'Yes' if pupil.laps else 'No', 'Yes' if pupil.service_child else 'No'] for idx, pupil in enumerate(pupils, start=1)]
        filters = dict(pupil_filters, class_id=class_id_raw, send=send_filter)
        if _pdf_requested():
            return _render_table_pdf('Pupil overview' if not anon_mode else 'Pupil overview — anonymised', headers, rows, filters, anon_mode, 'pupil_overview.pdf')
    return render_template(
        'admin/pupils.html',
        pupils=pupils,
        pupil_filters=pupil_filters,
        pupil_status_filter_choices=PUPIL_STATUS_FILTER_CHOICES,
        class_id_filter=class_id_raw,
        send_filter=send_filter,
        class_options=demo_filter_classes(SchoolClass.query.filter(SchoolClass.is_active.is_(True))).order_by(SchoolClass.year_group, SchoolClass.name).all(),
    )




def _selected_school_id_for_admin_actions() -> int | None:
    return current_user.school_id if not current_user.is_executive_admin else current_school_id()


def _require_admin_school_context(message: str = 'Select a school before downloading data.'):
    if _selected_school_id_for_admin_actions() is not None:
        return None
    flash(message, 'warning')
    return redirect(url_for('admin.classes'))


def _resolve_admin_quick_add_class(class_id_raw: str):
    class_id = int(class_id_raw or 0)
    school_id = _selected_school_id_for_admin_actions()
    if class_id <= 0 or school_id is None:
        return None
    return demo_filter_classes(SchoolClass.query.filter_by(id=class_id, school_id=school_id, is_active=True)).first()


def _handle_admin_quick_add_form_redirect():
    school_class = _resolve_admin_quick_add_class(request.form.get('class_id', '0'))
    pupil, error = create_quick_add_pupil(
        school_class=school_class,
        first_name=request.form.get('first_name', ''),
        last_name=request.form.get('last_name', ''),
        gender=request.form.get('gender', ''),
        pupil_premium=request.form.get('pupil_premium') == 'on',
        laps=request.form.get('laps') == 'on',
        service_child=request.form.get('service_child') == 'on',
        send=request.form.get('send') == 'on',
        join_year_group_raw=request.form.get('join_year_group', ''),
        join_date_raw=request.form.get('join_date', ''),
    )
    if error:
        flash(error, 'danger')
    else:
        flash(f'Added {pupil.full_name}.', 'success')
    return redirect(url_for('admin.pupils'))


@admin_bp.route('/api/admin/pupils/quick-add', methods=['POST'])
@login_required
def admin_quick_add_api():
    if current_user.role not in {'school_admin', 'admin', 'executive_admin'}:
        return jsonify({'ok': False, 'error': 'Forbidden'}), 403
    source = request.json if request.is_json else request.form
    school_class = _resolve_admin_quick_add_class(source.get('class_id', '0'))
    if current_user.is_executive_admin and _selected_school_id_for_admin_actions() is None:
        return jsonify({'ok': False, 'error': 'Select a school before adding pupils.'}), 400
    pupil, error = create_quick_add_pupil(
        school_class=school_class,
        first_name=source.get('first_name', ''),
        last_name=source.get('last_name', ''),
        gender=source.get('gender', ''),
        pupil_premium=source.get('pupil_premium') in (True, 'true', 'on', '1', 1),
        laps=source.get('laps') in (True, 'true', 'on', '1', 1),
        service_child=source.get('service_child') in (True, 'true', 'on', '1', 1),
        send=source.get('send') in (True, 'true', 'on', '1', 1),
        join_year_group_raw=source.get('join_year_group', ''),
        join_date_raw=source.get('join_date', ''),
    )
    if error:
        return jsonify({'ok': False, 'error': error}), 400
    return jsonify({'ok': True, 'pupil_id': pupil.id, 'name': pupil.full_name})

@admin_bp.route('/archive/pupils')
@login_required
@admin_required
def archived_pupils():
    archived_rows = (
        demo_filter_pupils(Pupil.query)
        .filter(Pupil.is_archived.is_(True))
        .order_by(Pupil.archived_at.desc(), Pupil.last_name.asc(), Pupil.first_name.asc())
        .all()
    )
    return render_template('admin/archived_pupils.html', pupils=archived_rows)


@admin_bp.route('/archive/pupils/<int:pupil_id>/confirm-delete')
@login_required
@admin_required
def confirm_permanent_delete_pupil(pupil_id: int):
    pupil = demo_filter_pupils(Pupil.query).filter(Pupil.id == pupil_id, Pupil.is_archived.is_(True)).first_or_404()
    linked_counts = _linked_pupil_record_counts(pupil.id)
    return render_template('admin/confirm_delete_pupil.html', pupil=pupil, linked_counts=linked_counts)


@admin_bp.route('/archive/pupils/<int:pupil_id>/delete', methods=['POST'])
@login_required
@admin_required
def permanent_delete_pupil(pupil_id: int):
    if is_demo_user():
        flash('This action is disabled in Demo Mode.', 'warning')
        return redirect(url_for('admin.archived_pupils'))
    pupil = demo_filter_pupils(Pupil.query).filter(Pupil.id == pupil_id, Pupil.is_archived.is_(True)).first_or_404()
    if request.form.get('confirm_delete_text', '').strip() != 'DELETE':
        flash('Type DELETE to confirm permanent deletion.', 'danger')
        return redirect(url_for('admin.confirm_permanent_delete_pupil', pupil_id=pupil.id))

    _delete_pupil_linked_data(pupil)
    pupil_name = pupil.full_name
    pupil_id_value = pupil.id
    school_id = pupil.school_id
    db.session.delete(pupil)
    log_audit_event(
        action='pupil_permanently_deleted',
        target_type='pupil',
        target_id=pupil_id_value,
        school_id=school_id,
        details=f'name={pupil_name}',
    )
    db.session.commit()
    flash(f'{pupil_name} and linked records were permanently deleted.', 'success')
    return redirect(url_for('admin.archived_pupils'))


@admin_bp.route('/pupils/manage', methods=['POST'])
@login_required
@admin_required
def manage_pupil():
    pupil = demo_filter_pupils(Pupil.query).filter(Pupil.id == int(request.form.get('pupil_id', '0'))).first_or_404()
    action = request.form.get('action', '').strip()
    if is_demo_user() and action in {'archive', 'restore', 'delete'}:
        flash('This action is disabled in Demo Mode.', 'warning')
        return _pupil_action_redirect()
    linked_counts = _linked_pupil_record_counts(pupil.id)
    has_linked_data = any(linked_counts.values())

    try:
        if action == 'archive':
            pupil.is_active = False
            pupil.is_archived = True
            pupil.archived_at = datetime.now(timezone.utc)
            pupil.archived_by_user_id = current_user.id
            pupil.archive_reason = request.form.get('archive_reason', '').strip() or 'Archived by admin'
            db.session.add(pupil)
            log_audit_event(
                action='pupil_archived',
                target_type='pupil',
                target_id=pupil.id,
                school_id=pupil.school_id,
                details=f'reason={pupil.archive_reason}',
            )
            db.session.commit()
            flash(f'Archived {pupil.full_name}. They are now hidden from active lists.', 'success')
        elif action == 'restore':
            pupil.is_active = True
            pupil.is_archived = False
            pupil.archived_at = None
            pupil.archived_by_user_id = None
            pupil.archive_reason = None
            db.session.add(pupil)
            log_audit_event(
                action='pupil_restored',
                target_type='pupil',
                target_id=pupil.id,
                school_id=pupil.school_id,
            )
            db.session.commit()
            flash(f'Restored {pupil.full_name}. They are active again.', 'success')
        elif action == 'delete':
            flash('Permanent deletion is only available from Archived pupils after confirmation.', 'warning')
        elif action == 'update_profile':
            join_year_group_raw = request.form.get('join_year_group', '').strip()
            pupil.join_year_group = int(join_year_group_raw) if join_year_group_raw != '' else None
            if pupil.join_year_group is not None and (pupil.join_year_group < 0 or pupil.join_year_group > 6):
                raise ValueError('Year Joined School must be between Reception and Year 6.')
            join_date_raw = request.form.get('join_date', '').strip()
            pupil.join_date = date.fromisoformat(join_date_raw) if join_date_raw else None
            pupil.pupil_premium = request.form.get('pupil_premium') == 'on'
            pupil.laps = request.form.get('laps') == 'on'
            pupil.service_child = request.form.get('service_child') == 'on'
            pupil.send = request.form.get('send') == 'on'
            db.session.add(pupil)
            db.session.commit()
            flash(f'Updated profile fields for {pupil.full_name}.', 'success')
        else:
            flash('Unknown pupil action.', 'warning')
    except ValueError as exc:
        db.session.rollback()
        flash(f'Pupil action failed: {exc}', 'danger')
    return _pupil_action_redirect()


def _parse_setting_form(prefix: str = '') -> dict:
    suffix = f'_{prefix}' if prefix else ''
    below_threshold = float(request.form.get(f'below_are_threshold_percent{suffix}', '0') or 0)
    return {
        'year_group': int(request.form.get(f'year_group{suffix}', '0')),
        'subject': request.form.get(f'subject{suffix}', '').strip(),
        'term': request.form.get(f'term{suffix}', '').strip(),
        'paper_1_name': request.form.get(f'paper_1_name{suffix}', '').strip(),
        'paper_1_max': int(request.form.get(f'paper_1_max{suffix}', '0')),
        'paper_2_name': request.form.get(f'paper_2_name{suffix}', '').strip(),
        'paper_2_max': int(request.form.get(f'paper_2_max{suffix}', '0')),
        'combined_max': int(request.form.get(f'combined_max{suffix}', '0') or 0),
        'below_are_threshold_percent': below_threshold,
        'on_track_threshold_percent': below_threshold,
        'exceeding_threshold_percent': float(request.form.get(f'exceeding_threshold_percent{suffix}', '0') or 0),
    }


@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    form = AssessmentSettingForm()
    filter_year_group = request.args.get('year_group', '').strip()
    filter_subject = request.args.get('subject', '').strip()
    filter_term = request.args.get('term', '').strip()

    if request.method == 'POST':
        action = request.form.get('action', 'create')
        try:
            if action == 'set-active-academic-year':
                year = AcademicYear.query.get(int(request.form.get('academic_year_id', '0')))
                school_id = _selected_school_id_for_admin_actions()
                school = School.query.get(school_id) if school_id else None
                if not year:
                    flash('Academic year could not be found.', 'danger')
                elif not school:
                    flash('Select a school before setting its working year.', 'warning')
                else:
                    school.current_academic_year = year
                    db.session.commit()
                    flash(f'School working academic year set to {year.name}.', 'success')
                return redirect(url_for('admin.settings'))
            if action == 'generate-academic-years':
                created_years = generate_next_missing_academic_years()
                if created_years:
                    flash(f"Generated academic years: {', '.join(year.name for year in created_years)}.", 'success')
                else:
                    flash('Academic years are already up to date.', 'info')
                return redirect(url_for('admin.settings'))
            if action == 'create':
                payload = validate_setting_payload(_parse_setting_form())
                setting = get_or_create_assessment_setting(payload['year_group'], payload['subject'], payload['term'])
                update_assessment_setting(setting, payload)
                recalculate_subject_results_for_scope(setting.year_group, setting.subject, setting.term)
                db.session.commit()
                flash(f"Saved {format_subject_name(setting.subject)} {setting.term.title()} settings for Year {setting.year_group}.", 'success')
            else:
                setting_id = int(request.form.get('setting_id', '0'))
                setting = AssessmentSetting.query.get_or_404(setting_id)
                payload = validate_setting_payload(_parse_setting_form(prefix=str(setting.id)))
                existing = AssessmentSetting.query.filter_by(year_group=payload['year_group'], subject=payload['subject'], term=payload['term']).first()
                if existing and existing.id != setting.id:
                    raise AssessmentValidationError('A setting already exists for that year group, subject, and term.')
                update_assessment_setting(setting, payload)
                recalculate_subject_results_for_scope(setting.year_group, setting.subject, setting.term)
                db.session.commit()
                flash(f"Updated {format_subject_name(setting.subject)} {setting.term.title()} settings for Year {setting.year_group}.", 'success')
        except (ValueError, AssessmentValidationError) as exc:
            db.session.rollback()
            flash(f'Settings could not be saved: {exc}', 'danger')

    settings_query = AssessmentSetting.query
    if filter_year_group:
        settings_query = settings_query.filter(AssessmentSetting.year_group == int(filter_year_group))
    if filter_subject:
        settings_query = settings_query.filter(AssessmentSetting.subject == filter_subject)
    if filter_term:
        settings_query = settings_query.filter(AssessmentSetting.term == filter_term)

    settings = settings_query.order_by(AssessmentSetting.year_group, AssessmentSetting.subject, AssessmentSetting.term).all()

    if request.method == 'GET' and filter_year_group and filter_subject and filter_term:
        form.year_group.data = int(filter_year_group)
        form.subject.data = filter_subject
        form.term.data = filter_term
        setting = AssessmentSetting.query.filter_by(year_group=int(filter_year_group), subject=filter_subject, term=filter_term).first()
        if setting:
            form.paper_1_name.data = setting.paper_1_name
            form.paper_1_max.data = setting.paper_1_max
            form.paper_2_name.data = setting.paper_2_name
            form.paper_2_max.data = setting.paper_2_max
            form.combined_max.data = setting.combined_max
            form.below_are_threshold_percent.data = setting.below_are_threshold_percent
            form.on_track_threshold_percent.data = setting.on_track_threshold_percent
            form.exceeding_threshold_percent.data = setting.exceeding_threshold_percent
        elif filter_subject in CORE_SUBJECTS:
            defaults = get_setting_defaults(filter_subject)
            form.paper_1_name.data = defaults['paper_1_name']
            form.paper_1_max.data = defaults['paper_1_max']
            form.paper_2_name.data = defaults['paper_2_name']
            form.paper_2_max.data = defaults['paper_2_max']
            form.combined_max.data = defaults['combined_max']
            form.below_are_threshold_percent.data = defaults['below_are_threshold_percent']
            form.on_track_threshold_percent.data = defaults['on_track_threshold_percent']
            form.exceeding_threshold_percent.data = defaults['exceeding_threshold_percent']

    return render_template(
        'admin/settings.html',
        settings=settings,
        filter_year_group=filter_year_group,
        filter_subject=filter_subject,
        filter_term=filter_term,
        filter_subject_choices=[('', 'All subjects')] + [(subject, format_subject_name(subject)) for subject in CORE_SUBJECTS],
        filter_term_choices=[('', 'All terms')] + TERMS,
        form=form,
        terms=TERMS,
        selected_year=get_school_working_academic_year(_selected_school_id_for_admin_actions()),
        academic_year_options=build_academic_year_options(),
    )




@admin_bp.route('/api/settings/quick-save', methods=['POST'])
@admin_bp.route('/api/admin/settings/quick-save', methods=['POST'])
@login_required
@admin_required
def settings_quick_save():
    print("AUTOSAVE HIT")
    data = request.get_json(silent=True) or {}
    field = (data.get('field') or '').strip()
    field_aliases = {
        'working_towards_pct': 'below_are_threshold_percent',
        'exceeding_pct': 'exceeding_threshold_percent',
    }
    field = field_aliases.get(field, field)
    allowed = {'year_group', 'term', 'subject', 'paper_1_name', 'paper_1_max', 'paper_2_name', 'paper_2_max', 'combined_max', 'below_are_threshold_percent', 'exceeding_threshold_percent'}
    if field not in allowed:
        return {'ok': False, 'error': 'Field not allowed'}, 400

    selected_school_id = current_school_id()
    if current_user.is_executive_admin:
        if selected_school_id is None:
            return {'ok': False, 'error': 'Select a school before editing settings'}, 400
        school_id = selected_school_id
    else:
        school_id = current_user.school_id
        if school_id is None:
            return {'ok': False, 'error': 'Your account is not linked to a school'}, 403

    setting = None
    try:
        setting_id = int(data.get('record_id') or 0)
    except (TypeError, ValueError):
        setting_id = 0
    if setting_id > 0:
        setting = AssessmentSetting.query.get(setting_id)
    if setting is None:
        try:
            year_group = int(data.get('year_group'))
        except (TypeError, ValueError):
            return {'ok': False, 'error': 'Missing record_id or valid row identity'}, 400
        subject = (data.get('subject') or '').strip()
        term = (data.get('term') or '').strip()
        if not subject or not term:
            return {'ok': False, 'error': 'Missing subject/term for row identity'}, 400
        query = AssessmentSetting.query.filter_by(year_group=year_group, subject=subject, term=term)
        if school_id is not None:
            query = query.filter(AssessmentSetting.school_id == school_id)
        setting = query.first()
    if setting is None:
        return {'ok': False, 'error': 'Setting row not found'}, 404
    if setting.school_id != school_id:
        return {'ok': False, 'error': 'Forbidden for selected school context', 'setting_school_id': setting.school_id, 'selected_school_id': school_id}, 403

    raw_value = data.get('value')
    try:
        if field in {'paper_1_max', 'paper_2_max', 'combined_max'}:
            if raw_value in (None, ''):
                raise ValueError('Blank numeric value')
            value = float(raw_value)
            if not value.is_integer():
                raise ValueError('Score maxima must be whole numbers')
            value = int(value)
        elif field == 'year_group':
            if raw_value in (None, ''):
                raise ValueError('Blank numeric value')
            value = int(raw_value)
        elif field in {'below_are_threshold_percent', 'exceeding_threshold_percent'}:
            if raw_value in (None, ''):
                raise ValueError('Blank numeric value')
            value = float(raw_value)
        else:
            value = (raw_value or '').strip()
        setattr(setting, field, value)
        payload = validate_setting_payload({
            'year_group': setting.year_group, 'subject': setting.subject, 'term': setting.term,
            'paper_1_name': setting.paper_1_name, 'paper_1_max': setting.paper_1_max,
            'paper_2_name': setting.paper_2_name, 'paper_2_max': setting.paper_2_max,
            'combined_max': setting.combined_max, 'below_are_threshold_percent': setting.below_are_threshold_percent,
            'on_track_threshold_percent': setting.below_are_threshold_percent, 'exceeding_threshold_percent': setting.exceeding_threshold_percent,
        })
        update_assessment_setting(setting, payload)
        db.session.add(setting); db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return {'ok': False, 'error': str(exc) or 'Invalid value'}, 400
    except AssessmentValidationError as exc:
        db.session.rollback()
        return {'ok': False, 'error': str(exc) or 'Invalid value'}, 400
    return {'ok': True, 'message': 'Saved'}
@admin_bp.route('/interventions')
@login_required
@admin_required
def interventions():
    from app.models import Intervention

    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    year_group = request.args.get('year_group', '').strip()
    class_id = request.args.get('class_id', '').strip()
    subject = request.args.get('subject', '').strip()
    status = request.args.get('status', 'active').strip() or 'active'
    print_mode = request.args.get('print', '0') == '1'
    anon_mode = request.args.get('anon', '0') == '1'

    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        flash('Select a school before viewing interventions.', 'warning')
        return redirect(url_for('admin.classes'))
    query = Intervention.query.join(Intervention.pupil).filter(Pupil.school_id == school_id, Pupil.is_demo.is_(is_demo_user()), Intervention.is_demo.is_(is_demo_user()))
    query = query.filter(Intervention.academic_year == academic_year)
    query = build_intervention_filters(query, year_group=year_group, class_id=class_id, subject=subject, status=status)
    rows = query.order_by(Intervention.is_active.desc(), Pupil.last_name, Pupil.first_name).all()
    current_scores = {row.id: get_current_score_for_intervention(row) for row in rows}
    if _pdf_requested():
        headers = ['Pupil', 'Class', 'Subject', 'Focus', 'Status', 'Current score']
        table_rows = [[f'Pupil {idx}' if anon_mode else row.pupil.full_name, row.pupil.school_class.name if row.pupil.school_class else '', row.subject.title(), row.focus_area or '', 'Active' if row.is_active else 'Closed', current_scores.get(row.id) or '—'] for idx, row in enumerate(rows, start=1)]
        filters = {'academic_year': academic_year, 'year_group': year_group, 'class_id': class_id, 'subject': subject, 'status': status}
        return _render_table_pdf('Interventions' if not anon_mode else 'Interventions — anonymised', headers, table_rows, filters, anon_mode, 'interventions.pdf')

    return render_template(
        'admin/interventions.html',
        interventions=rows,
        academic_year=academic_year,
        year_group=year_group,
        class_id=class_id,
        subject=subject,
        status=status,
        print_mode=print_mode,
        anon_mode=anon_mode,
        current_scores=current_scores,
        class_options=demo_filter_classes(SchoolClass.query.filter_by(is_active=True)).order_by(SchoolClass.year_group, SchoolClass.name).all(),
        subjects=CORE_SUBJECTS,
    )


@admin_bp.route('/sats', methods=['GET', 'POST'])
@login_required
@admin_required
def sats():
    school_id = current_school_id()
    return redirect(url_for('dashboards.sats_simple', school_id=school_id) if school_id is not None else url_for('dashboards.sats_simple'))

# legacy disabled
def _legacy_admin_sats_disabled():
    academic_year = request.values.get('academic_year', get_selected_current_academic_year())
    selected_class_id = request.values.get('class_id', '').strip()
    selected_tab_id_raw = request.values.get('exam_tab_id', '').strip()
    selected_class_id_int = int(selected_class_id) if selected_class_id.isdigit() else None
    if selected_class_id and selected_class_id_int is None:
        return _redirect_non_year6_sats_access()
    if selected_class_id_int and not _is_active_year6_class(selected_class_id_int):
        return _redirect_non_year6_sats_access()

    if request.method == 'POST':
        action = request.form.get('action', 'update_mode')
        try:
            if action == 'update_mode':
                set_tracker_mode(6, request.form.get('tracker_mode', 'sats'))
                flash(f'Year 6 tracker mode changed to {get_tracker_mode_label(6)}.', 'success')
            elif action == 'save_tab':
                tab_id = int(request.form.get('tab_id', '0')) or None
                tab = save_sats_tab({
                    'year_group': 6,
                    'name': request.form.get('tab_name', ''),
                    'display_order': request.form.get('tab_display_order', '1'),
                    'is_active': request.form.get('tab_is_active') == 'on',
                }, tab_id=tab_id)
                selected_tab_id_raw = str(tab.id)
                flash('SATs exam tab saved.', 'success')
            elif action == 'toggle_tab':
                tab = toggle_sats_tab(int(request.form.get('tab_id', '0')))
                selected_tab_id_raw = str(tab.id)
                flash(f"{tab.name} is now {'shown' if tab.is_active else 'hidden'}.", 'success')
            elif action == 'save_column':
                column_id = int(request.form.get('column_id', '0')) or None
                exam_tab_id = int(request.form.get('exam_tab_id', '0') or selected_tab_id_raw or '0')
                save_sats_column(6, {
                    'name': request.form.get('name', ''),
                    'subject': request.form.get('subject', ''),
                    'score_type': request.form.get('score_type', 'paper'),
                    'max_marks': request.form.get('max_marks', '0'),
                    'pass_percentage': request.form.get('pass_percentage', '0'),
                    'display_order': request.form.get('display_order', '1'),
                    'is_active': request.form.get('is_active') == 'on',
                }, exam_tab_id=exam_tab_id, column_id=column_id)
                selected_tab_id_raw = str(exam_tab_id)
                flash('SATs column saved.', 'success')
            elif action == 'toggle_column':
                column = toggle_sats_column(int(request.form.get('column_id', '0')))
                selected_tab_id_raw = str(column.exam_tab_id)
                flash(f"{column.name} is now {'shown' if column.is_active else 'hidden'}.", 'success')
            db.session.commit()
            return redirect(url_for('admin.sats', academic_year=academic_year, class_id=selected_class_id_int or None, exam_tab_id=selected_tab_id_raw or None))
        except (ValueError, SatsColumnValidationError) as exc:
            db.session.rollback()
            flash(f'SATs changes could not be saved: {exc}', 'danger')

    overview = build_year6_sats_overview(
        academic_year,
        class_id=selected_class_id_int,
        exam_tab_id=int(selected_tab_id_raw) if selected_tab_id_raw else None,
    )
    return render_template(
        'admin/sats.html',
        academic_year=academic_year,
        tracker_mode=get_tracker_mode(6),
        tracker_mode_label=get_tracker_mode_label(6),
        tracker_mode_options=SATS_TRACKER_MODES,
        class_options=demo_filter_classes(SchoolClass.query.filter_by(year_group=6, is_active=True)).order_by(SchoolClass.name).all(),
        selected_class_id=selected_class_id_int,
        columns=overview['columns'],
        all_columns=get_sats_columns(6, exam_tab_id=overview['selected_tab'].id if overview.get('selected_tab') else None, active_only=False),
        tabs=overview['tabs'],
        selected_tab=overview['selected_tab'],
        rows=overview['rows'],
        overview=overview['class_summaries'][0]['subject_totals'] if len(overview['class_summaries']) == 1 else {},
        class_summaries=overview['class_summaries'],
        sats_subject_choices=SATS_COLUMN_SUBJECTS,
        sats_score_type_choices=SATS_SCORE_TYPES,
    )


@admin_bp.route('/promotion', methods=['GET', 'POST'])
@login_required
@admin_required
def promotion():
    effective_school_id = _selected_school_id_for_admin_actions()
    working_year = get_school_working_academic_year(effective_school_id)
    academic_year = working_year.name
    next_year = build_next_academic_year(academic_year)
    mapping_rows = get_promotion_mapping_options(effective_school_id) if effective_school_id else []
    if request.method == 'POST':
        if effective_school_id is None:
            flash('Select a school before running promotion.', 'warning')
            return redirect(url_for('admin.promotion', academic_year=academic_year))
        if is_demo_user():
            flash('This action is disabled in Demo Mode.', 'warning')
            return redirect(url_for('admin.promotion', academic_year=academic_year))
        action = request.form.get('action', 'snapshot')
        try:
            if action == 'snapshot':
                count = snapshot_pupil_history(academic_year, effective_school_id)
                ensure_academic_year(academic_year)
                db.session.commit()
                flash(f'Archived {count} pupil class history record(s) for {academic_year}.', 'success')
            elif action == 'promote':
                if request.form.get('confirm_promotion') != 'yes':
                    raise ValueError(f'Confirm promotion from {academic_year} into {next_year} before continuing.')
                class_mapping: dict[int, int | None] = {}
                for row in mapping_rows:
                    source_class = row['source_class']
                    destination_choices = row['destination_classes']
                    selected_value = (request.form.get(f'destination_{source_class.id}') or '').strip()
                    if destination_choices:
                        if not selected_value:
                            raise ValueError(f'Please choose a destination class for {source_class.name}.')
                        if not selected_value.isdigit():
                            raise ValueError(f'Invalid destination class selected for {source_class.name}.')
                        selected_id = int(selected_value)
                        valid_ids = {item.id for item in destination_choices}
                        if selected_id not in valid_ids:
                            raise ValueError(f'Invalid destination class selected for {source_class.name}.')
                        class_mapping[source_class.id] = selected_id
                    else:
                        class_mapping[source_class.id] = None
                outcome = promote_pupils_to_next_year(academic_year, effective_school_id, class_mapping=class_mapping)
                db.session.commit()
                flash(f"Promotion complete. Moved {outcome['moved']} pupil(s), marked {outcome['leavers']} Year 6 leavers, and set {outcome['target_year']} as this school's working year.", 'success')
            return redirect(url_for('admin.promotion', academic_year=academic_year))
        except ValueError as exc:
            db.session.rollback()
            flash(f'Promotion changes could not be saved: {exc}', 'danger')

    history_rows = get_history_rows(academic_year, effective_school_id)
    return render_template(
        'admin/promotion.html',
        academic_year=academic_year,
        next_year=next_year,
        history_rows=history_rows,
        mapping_rows=mapping_rows,
    )


def _log_operational_error(*, route_name: str, import_type: str, retry_attempted: bool) -> None:
    user_name = getattr(current_user, 'username', 'anonymous') if current_user.is_authenticated else 'anonymous'
    current_app.logger.exception(
        'OperationalError during %s (route=%s, user=%s, import_type=%s, retry_attempted=%s)',
        route_name,
        request.path,
        user_name,
        import_type,
        retry_attempted,
    )


def _safe_import_commit(*, route_name: str, import_type: str, retry_attempted: bool = False) -> None:
    try:
        db.session.commit()
    except OperationalError:
        _log_operational_error(route_name=route_name, import_type=import_type, retry_attempted=retry_attempted)
        db.session.rollback()
        db.session.remove()
        raise


def _run_import_with_single_retry(import_type: str, rows: list[dict]):
    importers = {
        'combined': import_combined_results,
        'reception': import_reception_tracker,
        'sats_tracker': import_sats_tracker_results,
    }
    if import_type not in importers:
        raise CsvImportError('Unknown import type.')

    try:
        summary = importers[import_type](rows)
        _safe_import_commit(route_name='admin.imports', import_type=import_type)
        return summary, False
    except OperationalError:
        db.session.rollback()
        db.session.remove()
        summary = importers[import_type](rows)
        _safe_import_commit(route_name='admin.imports', import_type=import_type, retry_attempted=True)
        return summary, True



@admin_bp.route('/imports', methods=['GET', 'POST'])
@login_required
@admin_required
def imports():
    summary = None
    selected_import_type = 'combined'
    if request.method == 'POST':
        selected_import_type = request.form.get('import_type', 'combined')
        try:
            rows = parse_uploaded_csv(request.files.get('csv_file'))
            selected_year_record = AcademicYear.query.filter_by(id=request.form.get('academic_year_id')).first() if request.form.get('academic_year_id') else None
            fallback_year = selected_year_record.name if selected_year_record else get_selected_current_academic_year()
            for row in rows:
                if not (row.get('academic_year') or '').strip():
                    row['academic_year'] = fallback_year
            summary = import_combined_results(rows) if selected_import_type == 'combined' else _run_import_with_single_retry(selected_import_type, rows)[0]
            if request.form.get('confirm_save') == '1':
                _safe_import_commit(route_name='admin.imports', import_type=selected_import_type)
                action_label = 'Import finished'
            else:
                db.session.rollback()
                action_label = 'Preview finished (nothing saved)'
            if summary.errors:
                for error in summary.errors[:20]:
                    flash(error, 'warning')
            flash(
                f'{action_label}: total rows {summary.rows_processed}, pupils to create/created {summary.pupils_created}, '
                f'pupils to update/updated {summary.pupils_updated}, pupils matched {summary.pupils_matched}, '
                f'assessment results to import/imported {summary.subject_results_created + summary.subject_results_updated + summary.writing_results_created + summary.writing_results_updated + summary.tracker_entries_created + summary.tracker_entries_updated}, '
                f'rows skipped {summary.rows_skipped}, warnings/errors {summary.validation_errors}.',
                'success',
            )
        except CsvImportError as exc:
            db.session.rollback()
            flash(f'Import failed: {exc}', 'danger')

    overview = {
        'teachers': school_scoped_query(User, User.query.filter_by(role='teacher', is_demo=current_user.is_demo)).count(),
        'classes': demo_filter_classes(SchoolClass.query).count(),
        'pupils': demo_filter_pupils(Pupil.query).count(),
    }
    workbook_preview = session.pop('workbook_import_preview', None)
    ensure_default_academic_years()
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    current_year = get_selected_current_academic_year()
    selected_year_id = request.args.get('academic_year_id') or request.form.get('academic_year_id')
    if selected_year_id is None:
        current_year_record = next((year for year in years if year.name == current_year), None)
        selected_year_id = str(current_year_record.id) if current_year_record else None
    selected_year = next((year for year in years if str(year.id) == str(selected_year_id)), None)
    selected_year_label = selected_year.name if selected_year else current_year
    return render_template(
        'admin/imports.html',
        overview=overview,
        class_options=demo_filter_classes(SchoolClass.query.filter_by(is_active=True)).order_by(SchoolClass.year_group, SchoolClass.name).all(),
        current_year=current_year,
        years=years,
        selected_year_id=selected_year_id,
        selected_year=selected_year_label,
        reception_tracking_points=RECEPTION_TRACKING_POINTS,
        sats_tabs=get_sats_exam_tabs(6, include_inactive=False),
        summary=summary,
        selected_import_type=selected_import_type,
        workbook_preview=workbook_preview,
    )


@admin_bp.route('/imports/template/<template_type>')
@login_required
@admin_required
def download_import_template(template_type: str):
    template_map = {'combined', 'reception', 'sats_tracker'}
    if template_type not in template_map:
        flash('Unknown template type.', 'warning')
        return redirect(url_for('admin.imports'))
    csv_text = generate_csv(template_type)
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={"class_compass_combined_import_template.csv" if template_type == "combined" else f"{template_type}_template.csv"}' })


@admin_bp.route('/imports/full-template.xlsx')
@login_required
@admin_required
def download_full_template_xlsx():
    effective_school_id = _workbook_effective_school_id()
    if effective_school_id is None:
        flash('Select a school before downloading the full-school import workbook.', 'warning')
        return redirect(url_for('admin.imports'))
    wb = _build_full_template_workbook(effective_school_id)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='full_school_import_template.xlsx')


@admin_bp.route('/imports/full-workbook', methods=['POST'])
@login_required
@admin_required
def import_full_workbook():
    file = request.files.get('workbook_file')
    if not file or not file.filename.lower().endswith('.xlsx'):
        flash('Please upload a .xlsx workbook.', 'danger')
        return redirect(url_for('admin.imports'))
    wb = load_workbook(file, data_only=True)
    start_time = time.perf_counter()
    preview_errors = []
    preview_table = []
    created = 0
    updated = 0
    seen_pupil_rows = set()
    valid_genders = {'male', 'female', 'm', 'f', ''}
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        flash('Select a school before importing workbook data.', 'warning')
        return redirect(url_for('admin.imports'))
    selected_year_id = (request.form.get('academic_year_id') or '').strip()
    selected_year = AcademicYear.query.filter_by(id=selected_year_id).first() if selected_year_id else None
    selected_academic_year = (selected_year.name if selected_year else '').strip()
    academic_year = selected_academic_year or get_selected_current_academic_year()
    current_academic_year = get_selected_current_academic_year()
    batch_size = 200
    processed_rows = 0

    class_rows = demo_filter_classes(SchoolClass.query.options(joinedload(SchoolClass.teacher)).filter_by(school_id=school_id)).all()
    class_by_name = {_norm_key(c.name): c for c in class_rows}
    pupil_rows = demo_filter_pupils(Pupil.query.options(joinedload(Pupil.school_class)).filter_by(school_id=school_id)).all()
    pupil_by_lookup = {
        (_norm_key(p.first_name), _norm_key(p.last_name), p.class_id): p
        for p in pupil_rows
    }
    pupil_by_name = {
        (_norm_key(p.first_name), _norm_key(p.last_name)): p
        for p in pupil_rows
    }
    pupil_by_class_and_name = {
        (_norm_key(p.school_class.name) if p.school_class else '', _norm_key(_pupil_display_name(p))): p
        for p in pupil_rows
        if p.school_class
    }
    pupil_by_id = {p.id: p for p in pupil_rows}

    sheet_header_rows: dict[str, int] = {}

    def _headers_for(sheet_name: str) -> dict[str, int]:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        for header_row in range(1, min(ws.max_row, 5) + 1):
            headers = {_norm_key(cell.value): idx for idx, cell in enumerate(ws[header_row]) if _norm(cell.value)}
            if 'pupil_id' in headers or 'pupil' in headers or 'pupil_name' in headers or 'first_name' in headers:
                sheet_header_rows[sheet_name] = header_row
                return headers
        sheet_header_rows[sheet_name] = 1
        return {_norm_key(cell.value): idx for idx, cell in enumerate(ws[sheet_header_rows[sheet_name]]) if _norm(cell.value)}

    sheet_headers = {sheet_name: _headers_for(sheet_name) for sheet_name in wb.sheetnames}

    def _data_start_row(sheet_name: str) -> int:
        return sheet_header_rows.get(sheet_name, 1) + 1

    def _value(row, headers: dict[str, int], *names: str, default_index: int | None = None):
        for name in names:
            idx = headers.get(_norm_key(name))
            if idx is not None and idx < len(row):
                return row[idx]
        if default_index is not None and default_index < len(row):
            return row[default_index]
        return None

    def _pupil_from_assessment_row(row, headers: dict[str, int], default_class_index: int = 0, default_name_index: int = 1):
        pupil_id_raw = _value(row, headers, 'pupil_id')
        try:
            pupil_id = int(pupil_id_raw) if pupil_id_raw not in (None, '') else None
        except (TypeError, ValueError):
            pupil_id = None
        if pupil_id is not None and pupil_id in pupil_by_id:
            return pupil_by_id[pupil_id]
        class_name = _value(row, headers, 'class', default_index=default_class_index)
        pupil_name = _value(row, headers, 'pupil', 'pupil_name', 'full_name', default_index=default_name_index)
        return pupil_by_class_and_name.get((_norm_key(class_name), _norm_key(pupil_name)))

    subject_results = SubjectResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    subject_result_map = {(r.pupil_id, r.term, r.subject): r for r in subject_results}
    writing_results = WritingResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    writing_result_map = {(r.pupil_id, r.term): r for r in writing_results}
    foundation_results = FoundationResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    foundation_result_map = {(r.pupil_id, r.term, r.subject): r for r in foundation_results}
    reception_results = ReceptionTrackerEntry.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    reception_result_map = {(r.pupil_id, r.tracking_point, r.area_key): r for r in reception_results}
    sats_results = SatsResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    sats_result_map = {(r.pupil_id, r.exam_number): r for r in sats_results}
    sats_column_results = SatsColumnResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    sats_column_result_map = {(r.pupil_id, r.column_id): r for r in sats_column_results}
    sats_writing_results = SatsWritingResult.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    sats_writing_result_map = {(r.pupil_id, r.assessment_point): r for r in sats_writing_results}
    phonics_results = PhonicsScore.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    phonics_result_map = {(r.pupil_id, r.phonics_test_column_id): r for r in phonics_results}
    times_table_results = TimesTableScore.query.filter_by(school_id=school_id, academic_year=academic_year).all()
    times_table_result_map = {(r.pupil_id, r.times_table_test_column_id): r for r in times_table_results}
    history_rows = PupilClassHistory.query.filter(
        PupilClassHistory.school_id == school_id,
        PupilClassHistory.academic_year == academic_year,
        PupilClassHistory.pupil_id.in_([pupil.id for pupil in pupil_rows] or [0]),
    ).all()
    history_result_map = {row.pupil_id: row for row in history_rows}

    def _refresh_historical_lookup() -> None:
        pupil_by_class_and_name.clear()
        for pupil in pupil_rows:
            history_row = history_result_map.get(pupil.id)
            if history_row is not None:
                class_key = _norm_key(history_row.class_name)
            elif pupil.school_class is not None:
                class_key = _norm_key(pupil.school_class.name)
            else:
                continue
            pupil_by_class_and_name[(class_key, _norm_key(_pupil_display_name(pupil)))] = pupil

    if academic_year != current_academic_year:
        _refresh_historical_lookup()
    sats_tabs = get_sats_exam_tabs(6, include_inactive=True) if 'SATs' in wb.sheetnames else []
    sats_tabs_by_name = {_norm(tab.name).lower(): tab for tab in sats_tabs}
    sats_column_maps_by_tab = {
        tab.id: {_norm(column.column_key): column for column in get_sats_columns(6, exam_tab_id=tab.id, active_only=False) if column.column_key}
        for tab in sats_tabs
    }

    phonics_columns = PhonicsTestColumn.query.filter_by(school_id=school_id).all()
    phonics_column_map: dict[tuple[int, str], PhonicsTestColumn] = {}
    phonics_display_order_by_year: dict[int, int] = {}
    for column in phonics_columns:
        phonics_column_map[(column.year_group, _norm_key(column.name))] = column
        phonics_display_order_by_year[column.year_group] = max(phonics_display_order_by_year.get(column.year_group, 0), column.display_order or 0)

    def _flush_batch(sheet_name: str, row_idx: int) -> None:
        nonlocal processed_rows
        processed_rows += 1
        if processed_rows % batch_size == 0:
            db.session.flush()
            elapsed = time.perf_counter() - start_time
            current_app.logger.info('Importing %s sheet: %s rows processed (elapsed %.2fs)', sheet_name, processed_rows, elapsed)

    def _preview_row(sheet_name: str, row_number: int, status: str, detail: str, fix: str = ''):
        preview_table.append({'sheet': sheet_name, 'row_number': row_number, 'status': status, 'detail': detail, 'fix': fix})
    pupil_headers = sheet_headers.get('Pupils', {})
    for row_idx, row in enumerate(wb['Pupils'].iter_rows(min_row=_data_start_row('Pupils'), values_only=True), start=_data_start_row('Pupils')) if 'Pupils' in wb.sheetnames else []:
        pupil_id_raw = _value(row, pupil_headers, 'pupil_id')
        pupil_name = _norm(_value(row, pupil_headers, 'pupil', 'pupil_name', 'full_name', default_index=1))
        class_name = _norm(_value(row, pupil_headers, 'class', default_index=2))
        year_group_value = _value(row, pupil_headers, 'year_group', 'join_year_group', default_index=3)
        gender_value = _value(row, pupil_headers, 'gender', default_index=4)
        pp_value = _value(row, pupil_headers, 'pp', default_index=5)
        send_value = _value(row, pupil_headers, 'send', default_index=6)
        laps_value = _value(row, pupil_headers, 'laps', default_index=7)
        service_value = _value(row, pupil_headers, 'service_child', 'service', default_index=8)
        if not pupil_name or not class_name:
            _preview_row('Pupils', row_idx, 'skipped', 'Missing pupil name or class.', 'Complete the Pupil and Class columns.')
            continue
        row_key = (_norm_key(pupil_name), _norm_key(class_name))
        if row_key in seen_pupil_rows:
            preview_errors.append(f'Pupils row {row_idx}: duplicate row for {pupil_name} in {class_name}')
            _preview_row('Pupils', row_idx, 'warning', 'Duplicate row detected.', 'Remove duplicate rows so each pupil appears once.')
            continue
        seen_pupil_rows.add(row_key)
        gender_raw = _norm_key(gender_value)
        if gender_raw not in valid_genders:
            preview_errors.append(f'Pupils row {row_idx}: invalid gender "{_norm(gender_value)}"')
            _preview_row('Pupils', row_idx, 'warning', f'Invalid gender "{_norm(gender_value)}".', 'Use Male/Female or M/F.')
            continue
        school_class = class_by_name.get(_norm_key(class_name))
        if not school_class:
            preview_errors.append(f'Pupils row {row_idx}: class not found {class_name}')
            _preview_row('Pupils', row_idx, 'warning', f'Class "{class_name}" does not exist.', 'Create the class first or fix the class name.')
            continue
        try:
            pupil_id = int(pupil_id_raw) if pupil_id_raw not in (None, '') else None
        except (TypeError, ValueError):
            pupil_id = None
        pupil = pupil_by_id.get(pupil_id) if pupil_id is not None else None
        if not pupil:
            pupil = pupil_by_class_and_name.get((_norm_key(school_class.name), _norm_key(pupil_name)))
        first_name, last_name = _split_name(pupil_name)
        if not first_name:
            _preview_row('Pupils', row_idx, 'skipped', 'Missing pupil name.', 'Complete the Pupil column.')
            continue
        if not pupil:
            pupil = Pupil(first_name=first_name, last_name=last_name, class_id=school_class.id, gender=normalize_gender(_norm(gender_value)) or '', school_id=school_id, is_demo=current_user.is_demo)
            db.session.add(pupil)
            db.session.flush()
            pupil_by_id[pupil.id] = pupil
            pupil_by_lookup[(_norm_key(pupil.first_name), _norm_key(pupil.last_name), school_class.id)] = pupil
            pupil_by_name[(_norm_key(pupil.first_name), _norm_key(pupil.last_name))] = pupil
            pupil_rows.append(pupil)
            created += 1
            _preview_row('Pupils', row_idx, 'new', f'Will create pupil {pupil_name}.')
        else:
            if academic_year == current_academic_year and pupil.class_id != school_class.id:
                pupil.class_id = school_class.id
            _preview_row('Pupils', row_idx, 'updated', f'Will update pupil {pupil_name}.')

        history_row = history_result_map.get(pupil.id)
        if not history_row:
            history_row = PupilClassHistory(
                school_id=school_id,
                pupil_id=pupil.id,
                academic_year=academic_year,
                class_name=school_class.name,
                year_group=school_class.year_group,
                teacher_username=school_class.teacher.username if school_class.teacher else None,
            )
            db.session.add(history_row)
            history_result_map[pupil.id] = history_row
        else:
            history_row.class_name = school_class.name
            history_row.year_group = school_class.year_group
            history_row.teacher_username = school_class.teacher.username if school_class.teacher else None
        if academic_year == current_academic_year:
            pupil.gender = normalize_gender(_norm(gender_value)) or pupil.gender or ''
            pupil.pupil_premium = _norm(pp_value).lower() in {'1', 'true', 'yes', 'y'}
            pupil.laps = _norm(laps_value).lower() in {'1', 'true', 'yes', 'y'}
            pupil.service_child = _norm(service_value).lower() in {'1', 'true', 'yes', 'y'}
            pupil.send = _norm(send_value).lower() in {'1', 'true', 'yes', 'y'}
            try:
                pupil.join_year_group = int(year_group_value) if year_group_value not in (None, '') else pupil.join_year_group
            except (TypeError, ValueError):
                pass
        pupil_by_class_and_name[(_norm_key(school_class.name), _norm_key(_pupil_display_name(pupil)))] = pupil
        updated += 1
        _flush_batch('Pupils', row_idx)
    for sheet, subject in [('Maths','maths'),('Reading','reading'),('SPaG','spag')]:
        if sheet not in wb.sheetnames: continue
        for row_idx, row in enumerate(wb[sheet].iter_rows(min_row=_data_start_row(sheet), values_only=True), start=_data_start_row(sheet)):
            headers = sheet_headers.get(sheet, {})
            pupil = _pupil_from_assessment_row(row, headers)
            pupil_name = _value(row, headers, 'pupil', 'pupil_name', 'full_name', default_index=1)
            if not pupil:
                preview_errors.append(f'{sheet} row {row_idx}: pupil not matched {_norm(pupil_name)}')
                _preview_row(sheet, row_idx, 'warning', f'Pupil "{_norm(pupil_name)}" not found.', 'Check pupil ID, class, and pupil name match.')
                continue
            term=_norm(_value(row, headers, 'term', default_index=2)).lower()
            if term not in {'autumn','spring','summer'}:
                preview_errors.append(f'{sheet} row {row_idx}: invalid term {term}')
                _preview_row(sheet, row_idx, 'warning', f'Invalid term "{_norm(row[2])}".', 'Use Autumn, Spring, or Summer.')
                continue
            r = subject_result_map.get((pupil.id, term, subject))
            if not r:
                r=SubjectResult(pupil_id=pupil.id, school_id=school_id, academic_year=academic_year, term=term, subject=subject)
                db.session.add(r)
                subject_result_map[(pupil.id, term, subject)] = r
            try:
                paper_1_raw = _value(row, headers, 'arithmetic', 'paper_1', 'spelling', default_index=3)
                paper_2_raw = _value(row, headers, 'reasoning', 'paper_2', 'grammar', default_index=4)
                r.paper_1_score=int(paper_1_raw) if paper_1_raw not in (None,'') else None
                r.paper_2_score=int(paper_2_raw) if paper_2_raw not in (None,'') else None
            except (TypeError, ValueError):
                preview_errors.append(f'{sheet} row {row_idx}: invalid score value.')
                _preview_row(sheet, row_idx, 'warning', 'Invalid score format.', 'Use numeric values only.')
                continue
            r.combined_score=(r.paper_1_score or 0)+(r.paper_2_score or 0) if r.paper_1_score is not None and r.paper_2_score is not None else None
            r.notes=_norm(_value(row, headers, 'notes', default_index=5))
            _preview_row(sheet, row_idx, 'updated', f'Will import {sheet} data for {_norm(pupil_name)}.')
            _flush_batch(sheet, row_idx)
    if 'Phonics' in wb.sheetnames:
        headers = sheet_headers.get('Phonics', {})
        for row_idx, row in enumerate(wb['Phonics'].iter_rows(min_row=_data_start_row('Phonics'), values_only=True), start=_data_start_row('Phonics')):
            pupil = _pupil_from_assessment_row(row, headers)
            pupil_name = _value(row, headers, 'pupil', 'pupil_name', 'full_name', default_index=1)
            if not pupil:
                preview_errors.append(f'Phonics row {row_idx}: pupil not matched {_norm(pupil_name)}')
                continue
            test_name = _norm(_value(row, headers, 'test_name', default_index=2))
            if not test_name:
                preview_errors.append(f'Phonics: missing test_name for {_norm(pupil_name)}')
                continue
            score_raw = _value(row, headers, 'score', default_index=3)
            if score_raw in (None, ''):
                preview_errors.append(f'Phonics: missing score for {_norm(pupil_name)} ({test_name})')
                continue
            try:
                score_value = int(score_raw)
            except (TypeError, ValueError):
                preview_errors.append(f'Phonics: invalid score "{_norm(score_raw)}" for {_norm(pupil_name)} ({test_name})')
                continue
            col_key = (pupil.school_class.year_group, _norm_key(test_name))
            column = phonics_column_map.get(col_key)
            is_new_column = False
            if not column:
                next_order = phonics_display_order_by_year.get(pupil.school_class.year_group, 0) + 1
                phonics_display_order_by_year[pupil.school_class.year_group] = next_order
                column = PhonicsTestColumn(school_id=school_id, year_group=pupil.school_class.year_group, name=_norm(test_name), display_order=next_order, is_active=True)
                db.session.add(column)
                db.session.flush()
                phonics_column_map[col_key] = column
                is_new_column = True
            if is_new_column:
                preview_errors.append(f'Phonics: new test column will be created ({column.name}, Year {column.year_group})')
            else:
                preview_errors.append(f'Phonics: existing test column used ({column.name}, Year {column.year_group})')
            result = phonics_result_map.get((pupil.id, column.id))
            if not result:
                result = PhonicsScore(
                    school_id=school_id,
                    pupil_id=pupil.id,
                    academic_year=academic_year,
                    phonics_test_column_id=column.id,
                )
                db.session.add(result)
                phonics_result_map[(pupil.id, column.id)] = result
            result.score = score_value
            _preview_row('Phonics', row_idx, 'updated', f'Will save score for {_norm(pupil_name)} ({column.name}).')
            _flush_batch('Phonics', row_idx)
    if 'SATs' in wb.sheetnames:
        headers = sheet_headers.get('SATs', {})
        for row in wb['SATs'].iter_rows(min_row=_data_start_row('SATs'), values_only=True):
            pupil = _pupil_from_assessment_row(row, headers)
            pupil_name = _value(row, headers, 'pupil', 'pupil_name', 'full_name', default_index=1)
            if not pupil:
                preview_errors.append(f'SATs: pupil not matched {_norm(pupil_name)}')
                continue
            pupil_year_group = pupil.school_class.year_group if pupil.school_class else pupil.join_year_group
            if pupil_year_group != 6:
                preview_errors.append(f'SATs: skipped non-Year 6 pupil {_norm(pupil_name)}')
                continue
            assessment_point = _norm(_value(row, headers, 'assessment_point', 'exam_number', default_index=2))
            if assessment_point not in SATS_ASSESSMENT_POINTS:
                preview_errors.append(f'SATs: invalid assessment_point "{assessment_point}" for {_norm(pupil_name)}')
                continue
            tab = sats_tabs_by_name.get(assessment_point.lower())
            if not tab:
                preview_errors.append(f'SATs: assessment tab missing for {assessment_point}')
                continue
            column_map = sats_column_maps_by_tab.get(tab.id, {})
            imported_values = []
            sats_import_aliases = {
                'reading_raw': ('reading_raw', 'reading_paper', 'reading'),
                'reading_scaled': ('reading_scaled', 'reading_scaled_score'),
                'maths_arithmetic_raw': ('maths_arithmetic_raw', 'arithmetic'),
                'maths_reasoning_raw': ('maths_reasoning_raw', 'reasoning_1', 'reasoning'),
                'maths_scaled': ('maths_scaled', 'maths_scaled_score'),
                'spag_grammar_raw': ('spag_grammar_raw', 'grammar'),
                'spag_spelling_raw': ('spag_spelling_raw', 'spelling'),
                'spag_scaled': ('spag_scaled', 'spag_scaled_score'),
            }
            for csv_key in SATS_FIXED_COLUMNS.keys():
                raw_value = _value(row, headers, *sats_import_aliases.get(csv_key, (csv_key,)))
                if raw_value in (None, ''):
                    continue
                column = column_map.get(csv_key)
                if not column:
                    preview_errors.append(f'SATs: fixed column missing "{csv_key}" ({assessment_point})')
                    continue
                try:
                    score_value = int(raw_value)
                except (TypeError, ValueError):
                    preview_errors.append(f'SATs: invalid value for {csv_key} on {_norm(pupil_name)}')
                    continue
                result = sats_column_result_map.get((pupil.id, column.id))
                action = 'will update existing row' if result else 'will create new row'
                if not result:
                    result = SatsColumnResult(school_id=school_id, pupil_id=pupil.id, academic_year=academic_year, column_id=column.id)
                    sats_column_result_map[(pupil.id, column.id)] = result
                result.raw_score = score_value
                db.session.add(result)
                imported_values.append(f'{csv_key}={score_value}')
                preview_errors.append(f'SATs: {_norm(pupil_name)} {assessment_point} {csv_key} -> {action}')

            writing_band_raw = _value(row, headers, 'writing_band') if 'writing_band' in headers else None
            band_value = _normalize_writing_band(writing_band_raw)
            if _norm(writing_band_raw) and not band_value:
                preview_errors.append(f'SATs: invalid writing_band "{_norm(writing_band_raw)}" for {_norm(pupil_name)}')
            elif band_value:
                ap_index = SATS_ASSESSMENT_POINTS.index(assessment_point) + 1
                writing_row = sats_writing_result_map.get((pupil.id, ap_index))
                action = 'will update existing row' if writing_row else 'will create new row'
                if not writing_row:
                    writing_row = SatsWritingResult(school_id=school_id, pupil_id=pupil.id, academic_year=academic_year, assessment_point=ap_index)
                    sats_writing_result_map[(pupil.id, ap_index)] = writing_row
                writing_row.band = band_value
                writing_row.notes = _norm(_value(row, headers, 'notes'))
                db.session.add(writing_row)
                preview_errors.append(f'SATs: {_norm(pupil_name)} {assessment_point} writing_band={_norm(writing_band_raw)} -> {action}')
            if imported_values:
                preview_errors.append(f'SATs: values to import {_norm(pupil_name)} {assessment_point}: {", ".join(imported_values)}')
    if request.form.get('confirm_save')!='1':
        db.session.rollback()
        session['workbook_import_preview'] = {
            'academic_year': academic_year,
            'rows': preview_table[:120],
            'created': created,
            'updated': updated,
            'skipped': len([row for row in preview_table if row['status'] == 'skipped']),
            'warnings': len([row for row in preview_table if row['status'] == 'warning']),
        }
        for e in preview_errors[:20]: flash(e,'warning')
        flash(f'Preview complete. {created} pupils would be created/updated. Re-upload and click Save Workbook Import to apply.', 'info')
        return redirect(url_for('admin.imports'))
    try:
        db.session.flush()
        db.session.commit()
    except OperationalError:
        _log_operational_error(route_name='admin.import_full_workbook', import_type='full_workbook', retry_attempted=False)
        db.session.rollback()
        db.session.remove()
        raise
    ensure_academic_year(academic_year, mark_current=False)
    log_audit_event('import_full_workbook', 'school', school_id, school_id=school_id, details=f'academic_year={academic_year};created={created};updated={updated};warnings={len(preview_errors)}')
    for subject in ('maths', 'reading', 'spag'):
        for term in ('autumn', 'spring', 'summer'):
            recalculate_subject_results_for_scope(6, subject, term, academic_year=academic_year)
    for e in preview_errors[:20]: flash(e,'warning')
    flash('Workbook import complete.', 'success')
    return redirect(url_for('admin.imports'))


@admin_bp.route('/imports/full-workbook/export.xlsx')
@login_required
@admin_required
def export_full_workbook():
    effective_school_id = _workbook_effective_school_id()
    if effective_school_id is None:
        flash('Select a school before downloading the full-school import workbook.', 'warning')
        return redirect(url_for('admin.imports'))
    wb = _build_full_template_workbook(effective_school_id)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='full_school_import_template.xlsx')


@admin_bp.route('/reports/headline')
@login_required
@admin_required
def headline_report():
    context_redirect = _require_admin_school_context('Select a school before viewing the headline report.')
    if context_redirect:
        return context_redirect
    effective_school_id = _selected_school_id_for_admin_actions()
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    subject = (request.args.get('subject', 'maths') or 'maths').strip().lower()
    tracker_key = (request.args.get('tracker_key', '') or '').strip()
    year_group_raw = request.args.get('year_group', '').strip()
    year_group = int(year_group_raw) if year_group_raw.isdigit() else None
    if subject == 'eyfs' and year_group is None:
        year_group_raw = '0'
        year_group = 0
    if subject == 'times_tables' and year_group is None:
        year_group_raw = '4'
        year_group = 4
    if subject == 'sats' and year_group is None:
        year_group_raw = '6'
        year_group = 6
    send_filter = (request.args.get('send', 'all') or 'all').strip().lower()
    pupil_filters = build_admin_pupil_filter_state(request.args)
    pupil_filters['send'] = send_filter
    report = build_headline_report(
        subject=subject,
        academic_year=academic_year,
        year_group=year_group,
        filters=pupil_filters,
        tracker_key=tracker_key or None,
        school_id=effective_school_id,
    )
    current_app.logger.info(
        'headline_report school_scope user_id=%s role=%s user_school_id=%s effective_school_id=%s pupils=%s results=%s',
        current_user.id,
        current_user.role,
        current_user.school_id,
        effective_school_id,
        report.get('debug', {}).get('pupil_count'),
        report.get('debug', {}).get('result_count'),
    )
    tracker_options = []
    if subject == 'eyfs':
        tracker_options = [{'value': '', 'label': 'All tracking points'}] + [
            {'value': key, 'label': label} for key, label in RECEPTION_TRACKING_POINTS
        ]
    elif subject == 'phonics':
        phonics_years = [year_group] if year_group in {1, 2} else [1, 2]
        for year in phonics_years:
            for column in ensure_phonics_columns(year, effective_school_id):
                tracker_options.append({'value': str(column.id), 'label': f'Year {year} · {column.name}'})
    elif subject == 'times_tables':
        for column in ensure_times_tables_columns(4, effective_school_id):
            tracker_options.append({'value': str(column.id), 'label': column.name})
    elif subject == 'sats':
        tracker_options = [{'value': '', 'label': 'Latest active exam tab'}] + [
            {'value': str(tab.id), 'label': tab.name} for tab in get_sats_exam_tabs(6, include_inactive=False)
        ]
    return render_template(
        'admin/headline_report.html',
        report=report,
        subject=subject,
        tracker_key=tracker_key,
        tracker_options=tracker_options,
        academic_year=academic_year,
        year_group=year_group_raw,
        send_filter=send_filter,
        pupil_filters=pupil_filters,
        subjects=['writing', 'reading', 'maths', 'spag', 'eyfs', 'phonics', 'times_tables', 'sats'],
                boolean_filter_choices=BOOLEAN_FILTER_CHOICES,
    )


@admin_bp.route('/reports/headline/export')
@login_required
@admin_required
def export_headline_report():
    context_redirect = _require_admin_school_context('Select a school before downloading the headline report.')
    if context_redirect:
        return context_redirect
    effective_school_id = _selected_school_id_for_admin_actions()
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    subject = (request.args.get('subject', 'maths') or 'maths').strip().lower()
    tracker_key = (request.args.get('tracker_key', '') or '').strip()
    year_group_raw = request.args.get('year_group', '').strip()
    year_group = int(year_group_raw) if year_group_raw.isdigit() else None
    if subject == 'eyfs' and year_group is None:
        year_group = 0
    if subject == 'times_tables' and year_group is None:
        year_group = 4
    if subject == 'sats' and year_group is None:
        year_group = 6
    send_filter = (request.args.get('send', 'all') or 'all').strip().lower()
    pupil_filters = build_admin_pupil_filter_state(request.args)
    pupil_filters['send'] = send_filter
    report = build_headline_report(
        subject=subject,
        academic_year=academic_year,
        year_group=year_group,
        filters=pupil_filters,
        tracker_key=tracker_key or None,
        school_id=effective_school_id,
    )
    current_app.logger.info(
        'headline_report_export school_scope user_id=%s role=%s user_school_id=%s effective_school_id=%s pupils=%s results=%s',
        current_user.id,
        current_user.role,
        current_user.school_id,
        effective_school_id,
        report.get('debug', {}).get('pupil_count'),
        report.get('debug', {}).get('result_count'),
    )

    header = [report.get('row_header_label', 'Year group')]
    for term in report['buckets']:
        term_label = report['bucket_labels'][term]
        for measure_key in report['measure_keys']:
            header.append(f"{term_label} {report['measure_labels'][measure_key]}")
    table_rows = []
    for row in report['rows']:
        row_data = [row.get('label') or f"Year {row.get('year_group', '')}".strip()]
        for term in report['buckets']:
            for measure_key in report['measure_keys']:
                row_data.append(row['cells'][term][measure_key]['display'])
        table_rows.append(row_data)
    total_row = ['Whole school']
    for term in report['buckets']:
        for measure_key in report['measure_keys']:
            total_row.append(report['totals'][term][measure_key]['display'])
    table_rows.append(total_row)
    if request.args.get('format', '').strip().lower() == 'pdf':
        filters = {'academic_year': academic_year, 'subject': report['subject_label'], 'year_group': f'Year {year_group}' if year_group is not None else 'Whole school', 'send': send_filter}
        return _render_table_pdf('Headline report', header, table_rows, filters, False, 'headline_report.pdf', report['subject_label'])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Headline report'])
    writer.writerow(['Subject', report['subject_label']])
    writer.writerow(['Academic year', academic_year])
    writer.writerow(['Year group', f"Year {year_group}" if year_group else 'Whole school'])
    writer.writerow(['SEND', {'all':'All','yes':'Yes','no':'No'}.get(send_filter, 'All')])
    writer.writerow([])
    header = [report.get('row_header_label', 'Year group')]
    for term in report['buckets']:
        term_label = report['bucket_labels'][term]
        for measure_key in report['measure_keys']:
            header.append(f"{term_label} {report['measure_labels'][measure_key]}")
    writer.writerow(header)
    for row in report['rows']:
        row_data = [row.get('label') or f"Year {row.get('year_group', '')}".strip()]
        for term in report['buckets']:
            for measure_key in report['measure_keys']:
                row_data.append(row['cells'][term][measure_key]['display'])
        writer.writerow(row_data)
    total_row = ['Whole school']
    for term in report['buckets']:
        for measure_key in report['measure_keys']:
            total_row.append(report['totals'][term][measure_key]['display'])
    writer.writerow(total_row)
    csv_text = output.getvalue()
    filename_subject = subject if subject in {'maths', 'reading', 'spag', 'writing', 'eyfs', 'phonics', 'times_tables', 'sats'} else 'headline'
    return Response(
        csv_text,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=headline_{filename_subject}_{academic_year.replace("/", "-")}.csv'},
    )




@admin_bp.route('/reports/export-centre')
@login_required
@admin_required
def reports_export_centre():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    class_options = demo_filter_classes(SchoolClass.query.filter_by(is_active=True)).order_by(SchoolClass.name).all()
    year_groups = sorted({c.year_group for c in class_options})
    return render_template('admin/reports_export_centre.html', class_options=class_options, year_groups=year_groups)


@admin_bp.route('/reports/class-overview')
@login_required
@admin_required
def report_class_overview():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    query, filters = _apply_common_report_filters(demo_filter_pupils(Pupil.query))
    pupils = query.order_by(Pupil.last_name, Pupil.first_name).all()
    latest = _latest_subject_map([p.id for p in pupils])
    rows = []
    for p in pupils:
        rows.append([p.full_name, p.school_class.name, normalize_gender(p.gender) or '', 'Yes' if p.send else 'No', 'Yes' if p.pupil_premium else 'No', latest.get((p.id,'reading'),'—'), latest.get((p.id,'maths'),'—'), latest.get((p.id,'writing'),'—')])
    headers = ['Pupil','Class','Gender','SEND','PP','Reading','Maths','Writing']
    fmt = request.args.get('format','html')
    if fmt == 'csv':
        out=io.StringIO(); w=csv.writer(out); w.writerow(headers); w.writerows(rows)
        return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment; filename=class_overview_report.csv'})
    if fmt == 'xlsx':
        data = _build_xlsx(headers, rows, 'Class overview')
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=class_overview_report.xlsx'})
    if fmt == 'pdf':
        return _render_table_pdf('Class overview report', headers, rows, filters, False, 'class_overview_report.pdf', 'Printable class summary')
    return render_template('admin/report_table.html', title='Class overview report', subtitle='Printable class summary', headers=headers, rows=rows, filters=filters, anonymised=False)


@admin_bp.route('/reports/governor-summary')
@login_required
@admin_required
def report_governor_summary():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    query, filters = _apply_common_report_filters(demo_filter_pupils(Pupil.query))
    pupils = query.all()
    total = len(pupils) or 1
    def pct(n): return f"{round((n/total)*100,1)}%"
    rows = [
        ['PP vs non-PP', sum(1 for p in pupils if p.pupil_premium), sum(1 for p in pupils if not p.pupil_premium), pct(sum(1 for p in pupils if p.pupil_premium))],
        ['SEND vs non-SEND', sum(1 for p in pupils if p.send), sum(1 for p in pupils if not p.send), pct(sum(1 for p in pupils if p.send))],
        ['Gender (Female vs Male)', sum(1 for p in pupils if normalize_gender(p.gender)=='Female'), sum(1 for p in pupils if normalize_gender(p.gender)=='Male'), pct(sum(1 for p in pupils if normalize_gender(p.gender)=='Female'))],
    ]
    headers=['Cohort summary','Group A','Group B','Group A %']
    fmt=request.args.get('format','html')
    if fmt=='csv':
        out=io.StringIO();w=csv.writer(out);w.writerow(headers);w.writerows(rows)
        return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment; filename=governor_summary_report.csv'})
    if fmt=='xlsx':
        return Response(_build_xlsx(headers,rows,'Governor summary'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=governor_summary_report.xlsx'})
    if fmt == 'pdf':
        return _render_table_pdf('Governor / SLT anonymised summary', headers, rows, filters, True, 'governor_summary_report.pdf', 'No pupil names included')
    return render_template('admin/report_table.html', title='Governor / SLT anonymised summary', subtitle='No pupil names included', headers=headers, rows=rows, filters=filters, anonymised=True)
@admin_bp.route('/exports/subject-results')
@login_required
@admin_required
def export_subject_results():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    csv_text = export_subject_results_csv(
        class_id=int(request.args['class_id']) if request.args.get('class_id') else None,
        subject=request.args.get('subject') or None,
        academic_year=request.args.get('academic_year') or None,
        term=request.args.get('term') or None,
    )
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=subject_results_export.csv'})


@admin_bp.route('/exports/writing-results')
@login_required
@admin_required
def export_writing_results():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    csv_text = export_writing_results_csv(
        class_id=int(request.args['class_id']) if request.args.get('class_id') else None,
        academic_year=request.args.get('academic_year') or None,
        term=request.args.get('term') or None,
    )
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=writing_results_export.csv'})


@admin_bp.route('/exports/class-overview')
@login_required
@admin_required
def export_class_overview():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    csv_text = export_class_overview_csv(academic_year, class_id=int(request.args['class_id']) if request.args.get('class_id') else None)
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=class_overview_export.csv'})


@admin_bp.route('/exports/pupil-overview')
@login_required
@admin_required
def export_pupil_overview():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    csv_text = export_pupil_overview_csv(
        academic_year,
        class_id=int(request.args['class_id']) if request.args.get('class_id') else None,
        send=request.args.get('send', 'all'),
        anonymised=request.args.get('anon', '0') == '1',
    )
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=pupil_overview_export.csv'})


@admin_bp.route('/exports/sats')
@login_required
@admin_required
def export_sats():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    selected_class_id = int(request.args['class_id']) if request.args.get('class_id') else None
    if selected_class_id and not _is_active_year6_class(selected_class_id):
        return _redirect_non_year6_sats_access()
    csv_text = export_sats_results_csv(
        academic_year,
        class_id=selected_class_id,
        exam_tab_id=int(request.args['exam_tab_id']) if request.args.get('exam_tab_id') else None,
    )
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=sats_export.csv'})


@admin_bp.route('/exports/reception-tracker')
@login_required
@admin_required
def export_reception_tracker():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    tracking_point = request.args.get('tracking_point', RECEPTION_TRACKING_POINTS[0][0]).strip().lower()
    csv_text = export_reception_tracker_csv(academic_year, tracking_point)
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=reception_tracker_export.csv'})


@admin_bp.route('/exports/sats-tracker')
@login_required
@admin_required
def export_sats_tracker():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    exam_tab = request.args.get('exam_tab', '').strip()
    if not exam_tab:
        flash('Choose an exam tab for the SATs tracker export.', 'warning')
        return redirect(url_for('admin.imports'))
    try:
        csv_text = export_sats_tracker_csv(academic_year, exam_tab)
    except CsvImportError as exc:
        flash(str(exc), 'danger')
        return redirect(url_for('admin.imports'))
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=year6_sats_tracker_export.csv'})


@admin_bp.route('/exports/interventions')
@login_required
@admin_required
def export_interventions():
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    anon_mode = request.args.get('anon', '0') == '1'
    school_id = _selected_school_id_for_admin_actions()
    if school_id is None:
        flash('Select a school before exporting interventions.', 'warning')
        return redirect(url_for('admin.interventions', academic_year=academic_year))
    query = Intervention.query.join(Intervention.pupil).filter(Intervention.academic_year == academic_year, Pupil.school_id == school_id)
    if request.args.get('class_id'):
        query = query.filter(Pupil.class_id == int(request.args['class_id']))
    rows = query.order_by(Pupil.last_name, Pupil.first_name).all()
    current_scores = {row.id: get_current_score_for_intervention(row) for row in rows}
    csv_text = export_interventions_csv(
        academic_year,
        class_id=int(request.args['class_id']) if request.args.get('class_id') else None,
        anonymised=anon_mode,
        current_scores=current_scores,
    )
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=interventions_export.csv'})


@admin_bp.route('/exports/history')
@login_required
@admin_required
def export_history():
    context_redirect = _require_admin_school_context()
    if context_redirect:
        return context_redirect
    academic_year = request.args.get('academic_year', get_selected_current_academic_year())
    csv_text = export_history_csv(academic_year)
    return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=promotion_history_export.csv'})


@admin_bp.route('/data-quality')
@login_required
@admin_required
def data_quality():
    term_options = [('autumn', 'Autumn'), ('spring', 'Spring'), ('summer', 'Summer')]
    valid_terms = {value for value, _label in term_options}
    selected_term = (request.args.get('term') or 'summer').strip().lower()
    if selected_term not in valid_terms:
        selected_term = 'summer'
    selected_term_label = dict(term_options)[selected_term]

    academic_year = request.args.get('academic_year') or get_selected_current_academic_year()
    pupil_query = demo_filter_pupils(Pupil.query.filter_by(is_active=True))
    class_query = demo_filter_classes(SchoolClass.query.filter_by(is_active=True))

    issues = []
    issues.append({'issue': 'Pupils without class', 'count': pupil_query.filter(Pupil.class_id.is_(None)).count(), 'action_link': url_for('admin.pupils')})
    issues.append({'issue': 'Pupils missing gender', 'count': pupil_query.filter(func.coalesce(Pupil.gender, '') == '').count(), 'action_link': url_for('admin.pupils')})
    issues.append({'issue': 'Pupils missing PP/LAPS/Service/SEND values', 'count': pupil_query.filter(or_(Pupil.pupil_premium.is_(None), Pupil.laps.is_(None), Pupil.service_child.is_(None), Pupil.send.is_(None))).count(), 'action_link': url_for('admin.pupils')})
    issues.append({'issue': 'Pupils without join year group', 'count': pupil_query.filter(Pupil.join_year_group.is_(None)).count(), 'action_link': url_for('admin.pupils')})

    duplicates = (
        pupil_query.with_entities(Pupil.first_name, Pupil.last_name, Pupil.class_id, func.count(Pupil.id).label('dup_count'))
        .group_by(Pupil.first_name, Pupil.last_name, Pupil.class_id)
        .having(func.count(Pupil.id) > 1)
        .all()
    )
    issues.append({'issue': 'Duplicate pupil names in same class', 'count': len(duplicates), 'action_link': url_for('admin.pupils')})
    issues.append({'issue': 'Classes with no teacher', 'count': class_query.filter(SchoolClass.teacher_id.is_(None)).count(), 'action_link': url_for('admin.classes')})

    teachers_no_class = school_scoped_query(User, User.query).filter_by(role='teacher', is_active=True, is_demo=is_demo_user()).outerjoin(SchoolClass, and_(SchoolClass.teacher_id == User.id, SchoolClass.is_active.is_(True))).filter(SchoolClass.id.is_(None)).count()
    issues.append({'issue': 'Teachers with no class', 'count': teachers_no_class, 'action_link': url_for('admin.users')})

    active_pupil_count = pupil_query.count()
    assessed_ids = {
        pid
        for (pid,) in school_scoped_query(
            SubjectResult,
            SubjectResult.query.with_entities(SubjectResult.pupil_id)
            .filter(
                SubjectResult.academic_year == academic_year,
                SubjectResult.term.in_([selected_term, selected_term_label]),
            )
            .distinct(),
        ).all()
    }
    issues.append({'issue': f'Missing scores for {selected_term_label} {academic_year}', 'count': max(active_pupil_count - len(assessed_ids), 0), 'action_link': url_for('admin.classes', academic_year=academic_year)})

    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    recent_ids = {pid for (pid,) in school_scoped_query(SubjectResult, SubjectResult.query.with_entities(SubjectResult.pupil_id).filter(SubjectResult.updated_at >= thirty_days_ago).distinct()).all()}
    issues.append({'issue': 'Pupils with no recent assessment data (30 days)', 'count': max(active_pupil_count - len(recent_ids), 0), 'action_link': url_for('admin.classes')})

    live_archived = pupil_query.filter(Pupil.is_archived.is_(True)).count()
    issues.append({'issue': 'Archived pupils still in live tables', 'count': live_archived, 'action_link': url_for('admin.archived_pupils')})

    return render_template(
        'admin/data_quality.html',
        issues=issues,
        term_options=term_options,
        selected_term=selected_term,
        selected_term_label=selected_term_label,
        academic_year=academic_year,
    )


@admin_bp.route('/setup-checklist')
@login_required
@admin_required
def setup_checklist():
    context_redirect = _require_admin_school_context('Select a school before viewing the setup checklist.')
    if context_redirect:
        return context_redirect
    school_id = current_school_id()
    classes_count = demo_filter_classes(SchoolClass.query.filter_by(is_active=True)).count()
    teachers_count = school_scoped_query(User, User.query).filter_by(role='teacher', is_active=True, is_demo=is_demo_user()).count()
    assigned_teacher_count = demo_filter_classes(SchoolClass.query.filter(SchoolClass.teacher_id.is_not(None), SchoolClass.is_active.is_(True))).count()
    pupils_count = demo_filter_pupils(Pupil.query.filter_by(is_active=True)).count()
    settings_count = school_scoped_query(AssessmentSetting, AssessmentSetting.query).count()
    send_quality_ok = demo_filter_pupils(Pupil.query.filter_by(is_active=True)).filter(Pupil.send.is_(None)).count() == 0
    school = School.query.get(school_id) if school_id else None

    items = [
        {'label': 'School created', 'complete': school is not None, 'link': url_for('dashboards.index')},
        {'label': 'Classes created', 'complete': classes_count > 0, 'link': url_for('admin.classes')},
        {'label': 'Teachers created', 'complete': teachers_count > 0, 'link': url_for('admin.users')},
        {'label': 'Teachers assigned to classes', 'complete': assigned_teacher_count > 0, 'link': url_for('admin.classes')},
        {'label': 'Pupils uploaded', 'complete': pupils_count > 0, 'link': url_for('admin.imports')},
        {'label': 'Assessment settings completed', 'complete': settings_count > 0, 'link': url_for('admin.settings')},
        {'label': 'SEND/PP/LAPS/Service data checked', 'complete': send_quality_ok, 'link': url_for('admin.pupils')},
        {'label': 'Privacy/terms links active', 'complete': True, 'link': url_for('auth.login')},
        {'label': 'Demo data not mixed with real school', 'complete': school is None or (school.is_demo == is_demo_user()), 'link': url_for('admin.pupils')},
    ]
    return render_template('admin/setup_checklist.html', items=items)
