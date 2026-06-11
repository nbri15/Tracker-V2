"""Standalone Maths Fundamentals services and assessment engine."""

from __future__ import annotations

from collections import defaultdict
from base64 import b64encode
from io import BytesIO
from datetime import datetime, timedelta, timezone
import json
import random
import re
from statistics import mean
from typing import Any

from sqlalchemy import func

QUESTIONS_PER_LEVEL = 10
PASS_MARK = 8
MAX_LEVEL = 15
VISUAL_MARKER_RE = re.compile(r'\s*\[\[mf_visual:(dot_pattern|five_frame|ten_frame):(\d+)\]\]')
VARIANT_MARKER_RE = re.compile(r'\s*\[\[mf_variant:\d+\]\]')

from app.extensions import db
from app.models import (
    MathsFundamentalAttempt,
    MathsFundamentalQuestion,
    MathsFundamentalResult,
    MathsFundamentalSkill,
    MathsFundamentalStrand,
    MathsFundamentalsSession,
    MathsQuestionTemplate,
    Pupil,
    PupilQrToken,
    SchoolClass,
)
from app.services.assessments import apply_admin_pupil_filters

DEFAULT_STRANDS = [
    'Number Sense',
    'Number Bonds',
    'Addition & Subtraction',
    'Multiplication',
    'Division',
    'Place Value',
    'Fractions & Decimals',
]

WORKBOOK_STRAND_SHEETS = set(DEFAULT_STRANDS)

SHORT_STRAND_NAMES = {
    'Addition & Subtraction': 'Add/Sub',
    'Fractions & Decimals': 'Fractions',
}

COLUMN_ALIASES = {
    'strand': {'strand', 'strands'},
    'level': {'level', 'ladder level'},
    'band': {'band', 'stage'},
    'skill_text': {'skill', 'skill text', 'objective', 'learning objective', 'statement', 'fundamental'},
    'teaching_prompt': {'teaching idea', 'teaching prompt', 'teaching', 'teacher prompt', 'potential teaching / intervention'},
    'question_prompt': {'assessment question', 'question', 'question prompt', 'questions', 'question & app prompt ideas'},
    'question_type': {'question type', 'type'},
    'evidence': {'evidence', 'suggested evidence'},
    'notes': {'notes', 'note'},
    'template_text': {'template', 'question template', 'template text'},
    'generator_type': {'generator type', 'generator'},
    'generator_config_json': {'generator config', 'config', 'generator_config_json'},
    'answer_type': {'answer type', 'answer'},
    'difficulty': {'difficulty'},
}


def qr_code_data_uri(value: str) -> str:
    """Return a base64 PNG data URI for a QR code URL."""

    import qrcode

    qr = qrcode.QRCode(version=None, box_size=6, border=2)
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(fill_color='black', back_color='white')
    buffer = BytesIO()
    image.save(buffer, format='PNG')
    return 'data:image/png;base64,' + b64encode(buffer.getvalue()).decode('ascii')


def strand_short_name(strand: MathsFundamentalStrand) -> str:
    return SHORT_STRAND_NAMES.get(strand.name, strand.name)


def level_colour_class(level: int | None) -> str:
    level = int(level or 0)
    if level <= 0:
        return 'mf-empty'
    if level <= 4:
        return 'mf-red'
    if level <= 8:
        return 'mf-amber'
    if level <= 12:
        return 'mf-light-green'
    return 'mf-dark-green'


def ensure_default_ladder() -> None:
    """Create starter strands/skills only when no ladder has been imported yet."""

    if MathsFundamentalStrand.query.first():
        return
    for order, strand_name in enumerate(DEFAULT_STRANDS, start=1):
        strand = MathsFundamentalStrand(name=strand_name, display_order=order, is_active=True)
        db.session.add(strand)
        db.session.flush()
        for level in range(1, 16):
            skill = MathsFundamentalSkill(
                strand_id=strand.id,
                level=level,
                band=f'Level {level}',
                skill_text=f'{strand_name} Level {level}',
                teaching_prompt=f'Teach and rehearse {strand_name.lower()} at Level {level}.',
                question_prompt='Answer the generated question.',
                question_type='template',
                display_order=level,
            )
            db.session.add(skill)
            db.session.flush()
            db.session.add(MathsQuestionTemplate(
                skill_id=skill.id,
                generator_type='template',
                template_text='What is {a} + {b}?',
                generator_config_json=json.dumps({'a': [1, max(5, level * 2)], 'b': [1, max(5, level * 2)]}),
                answer_type='number',
                difficulty='standard',
                is_active=True,
            ))
    db.session.commit()


def active_strands() -> list[MathsFundamentalStrand]:
    ensure_default_ladder()
    return MathsFundamentalStrand.query.filter_by(is_active=True).order_by(MathsFundamentalStrand.display_order, MathsFundamentalStrand.name).all()


def get_or_create_qr_token(pupil: Pupil) -> PupilQrToken:
    token = PupilQrToken.query.filter_by(school_id=pupil.school_id, pupil_id=pupil.id).first()
    if token:
        return token
    token = PupilQrToken(school_id=pupil.school_id, pupil_id=pupil.id, token=PupilQrToken.generate_token())
    db.session.add(token)
    db.session.commit()
    return token


def current_results_map(pupils: list[Pupil], academic_year: str) -> dict[tuple[int, int], MathsFundamentalResult]:
    pupil_ids = [pupil.id for pupil in pupils]
    if not pupil_ids:
        return {}
    rows = MathsFundamentalResult.query.filter(
        MathsFundamentalResult.pupil_id.in_(pupil_ids),
        MathsFundamentalResult.academic_year == academic_year,
    ).all()
    return {(row.pupil_id, row.strand_id): row for row in rows}


def build_teacher_rows(pupils: list[Pupil], strands: list[MathsFundamentalStrand], academic_year: str) -> list[dict[str, Any]]:
    results = current_results_map(pupils, academic_year)
    rows = []
    for pupil in pupils:
        rows.append({
            'pupil': pupil,
            'qr_token': get_or_create_qr_token(pupil),
            'levels': {strand.id: results.get((pupil.id, strand.id)) for strand in strands},
        })
    return rows


def normalize_header(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_header(value) for value in headers]
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = index
                break
    return mapping


def import_ladder_from_workbook(path: str) -> dict[str, int]:
    """Import strands, skills and templates from the workbook without touching pupil data."""

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    imported = {'strands': 0, 'skills': 0, 'templates': 0}

    for strand_position, worksheet in enumerate(workbook.worksheets, start=1):
        default_strand = worksheet.title.strip()
        if default_strand not in WORKBOOK_STRAND_SHEETS:
            continue
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = None
        mapping = {}
        for row_index, row in enumerate(rows[:10]):
            candidate = header_map(list(row))
            if 'level' in candidate and 'skill_text' in candidate:
                headers = row_index
                mapping = candidate
                break
        if headers is None:
            continue

        strand = MathsFundamentalStrand.query.filter(func.lower(MathsFundamentalStrand.name) == default_strand.lower()).first()
        if not strand:
            strand = MathsFundamentalStrand(name=default_strand)
            imported['strands'] += 1
        strand.display_order = strand_position
        strand.is_active = True
        db.session.add(strand)
        db.session.flush()

        imported_skill_ids = set()
        for display_order, row in enumerate(rows[headers + 1:], start=1):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            level_raw = row[mapping['level']]
            try:
                level = int(float(str(level_raw).strip()))
            except (TypeError, ValueError):
                continue

            skill_text = _cell(row, mapping, 'skill_text')
            if not skill_text:
                continue
            skill = MathsFundamentalSkill.query.filter_by(strand_id=strand.id, level=level, display_order=display_order).first()
            if not skill:
                skill = MathsFundamentalSkill(strand_id=strand.id, level=level, display_order=display_order, skill_text=skill_text)
                imported['skills'] += 1
            skill.band = _cell(row, mapping, 'band')
            skill.skill_text = skill_text
            skill.teaching_prompt = _cell(row, mapping, 'teaching_prompt')
            skill.question_prompt = _cell(row, mapping, 'question_prompt')
            skill.question_type = _cell(row, mapping, 'question_type')
            skill.evidence = _cell(row, mapping, 'evidence')
            skill.notes = _cell(row, mapping, 'notes')
            skill.display_order = display_order
            db.session.add(skill)
            db.session.flush()
            imported_skill_ids.add(skill.id)

            template = skill.templates.order_by(MathsQuestionTemplate.id).first()
            if not template:
                template = MathsQuestionTemplate(skill_id=skill.id)
                imported['templates'] += 1
            template.generator_type = _cell(row, mapping, 'generator_type') or infer_generator_type(default_strand, skill.skill_text, skill.question_prompt)
            template.template_text = _cell(row, mapping, 'template_text') or skill.question_prompt or skill.skill_text
            template.generator_config_json = _cell(row, mapping, 'generator_config_json') or json.dumps(default_generator_config(default_strand, skill.level, template.generator_type))
            template.answer_type = _cell(row, mapping, 'answer_type') or default_answer_type(template.generator_type)
            template.difficulty = _cell(row, mapping, 'difficulty') or 'standard'
            template.is_active = True
            db.session.add(template)

        stale_skills = MathsFundamentalSkill.query.filter(
            MathsFundamentalSkill.strand_id == strand.id,
            ~MathsFundamentalSkill.id.in_(imported_skill_ids),
            MathsFundamentalSkill.level < 900,
        ).all() if imported_skill_ids else []
        for stale_skill in stale_skills:
            stale_skill.level = 900 + stale_skill.level
            stale_skill.display_order = 900 + stale_skill.display_order
            for stale_template in stale_skill.templates.all():
                stale_template.is_active = False
                db.session.add(stale_template)
            db.session.add(stale_skill)
    db.session.commit()
    return imported


def infer_generator_type(strand_name: str, skill_text: str | None, question_prompt: str | None = None) -> str:
    """Map workbook skill wording to a conservative question generator."""

    skill_prompt_text = f"{skill_text or ''} {question_prompt or ''}".lower()
    text = f"{strand_name} {skill_prompt_text}".lower()
    if any(word in text for word in ('subitise', 'subitize', 'dot', 'five-frame', 'five frame', 'ten-frame', 'ten frame')):
        return 'subitising'
    if 'number bond' in text or 'bonds' in text:
        return 'number_bonds'
    if 'division' in text or 'divide' in text or 'sharing' in text or 'grouping' in text or '÷' in text:
        return 'division'
    if 'multiplication' in text or 'times table' in text or 'times tables' in text or 'multiply' in text or '×' in text:
        return 'multiplication'
    if 'subtraction' in skill_prompt_text or 'subtract' in skill_prompt_text or 'minus' in skill_prompt_text:
        return 'subtraction'
    if 'addition' in skill_prompt_text or 'add ' in skill_prompt_text or '+' in skill_prompt_text or strand_name == 'Addition & Subtraction':
        return 'addition'
    if 'count' in text or 'one more' in text or 'one less' in text:
        return 'counting'
    return 'template'


def default_answer_type(generator_type: str) -> str:
    return 'number' if generator_type in {'subitising', 'counting', 'number_bonds', 'addition', 'subtraction', 'multiplication', 'division'} else 'text'


def default_generator_config(strand_name: str, level: int, generator_type: str) -> dict[str, Any]:
    if generator_type == 'subitising':
        return {'n': [1, 5 if level <= 1 else 10]}
    if generator_type == 'counting':
        limit = 20 if level <= 5 else 100
        return {'n': [1, limit], 'limit': limit}
    if generator_type == 'number_bonds':
        target = 5 if level <= 1 else (10 if level <= 3 else 20)
        return {'target': target}
    if generator_type in {'multiplication', 'division'}:
        factor = min(12, max(2, level + 1))
        return {'a': [1, 12], 'b': [2, factor]}
    if generator_type in {'addition', 'subtraction'}:
        upper = min(100, max(10, level * 10))
        return {'a': [1, upper], 'b': [1, upper]}
    return {}


def _cell(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> str | None:
    index = mapping.get(field)
    if index is None or index >= len(row) or row[index] is None:
        return None
    value = str(row[index]).strip()
    return value or None


def generate_question(skill: MathsFundamentalSkill, pupil_id: int | None = None) -> MathsFundamentalQuestion:
    template = skill.templates.filter_by(is_active=True).order_by(MathsQuestionTemplate.difficulty, MathsQuestionTemplate.id).first()
    if not template:
        generator_type = infer_generator_type(skill.strand.name, skill.skill_text, skill.question_prompt)
        template = MathsQuestionTemplate(
            skill_id=skill.id,
            generator_type=generator_type,
            template_text=skill.question_prompt or skill.skill_text,
            generator_config_json=json.dumps(default_generator_config(skill.strand.name, skill.level, generator_type)),
            answer_type=default_answer_type(generator_type),
        )
        db.session.add(template)
        db.session.flush()

    question_text, correct_answer = render_generated_question(skill, template)
    teacher_mark_required = correct_answer is None
    return MathsFundamentalQuestion(
        skill_id=skill.id,
        question_text=question_text,
        correct_answer='' if correct_answer is None else str(correct_answer),
        teacher_mark_required=teacher_mark_required,
        level=skill.level,
    )


def render_generated_question(skill: MathsFundamentalSkill, template: MathsQuestionTemplate) -> tuple[str, int | float | None]:
    """Render a question from workbook prompts using strand-safe generator rules."""

    generator_type = template.generator_type or infer_generator_type(skill.strand.name, skill.skill_text, skill.question_prompt)
    config = _load_generator_config(template.generator_config_json)
    prompt = choose_prompt(template.template_text or skill.question_prompt or skill.skill_text)

    if skill.strand.name == 'Number Sense' and generator_type == 'template':
        generator_type = 'subitising'

    if generator_type == 'subitising':
        low, high = _bounds(config.get('n'), 1, 10)
        high = max(high, QUESTIONS_PER_LEVEL)
        n = random.randint(max(1, low), min(10, high))
        visual_type = 'five_frame' if n <= 5 else 'ten_frame'
        if skill.strand.name == 'Number Sense' and skill.level >= 4:
            visual_type = random.choice(['dot_pattern', visual_type])
        return f"How many dots do you see? {_visual_marker(visual_type, n)}", n
    if generator_type == 'counting':
        limit = int(config.get('limit') or (20 if skill.level <= 5 else 100))
        count_low = 1 if skill.strand.name == 'Number Sense' else 2
        n = random.randint(count_low, max(count_low, min(limit - 1, 10 if skill.strand.name == 'Number Sense' else limit - 1)))
        if skill.strand.name == 'Number Sense':
            visual_type = 'five_frame' if n <= 5 else 'ten_frame'
            if 'back' in skill.skill_text.lower() or 'before' in prompt.lower() or 'one less' in skill.skill_text.lower():
                return f"Look at the frame. What is one less? {_visual_marker(visual_type, n)}", n - 1
            return f"Look at the frame. What is one more? {_visual_marker(visual_type, n)}", n + 1
        if 'back' in skill.skill_text.lower() or 'before' in prompt.lower() or 'one less' in skill.skill_text.lower():
            return f"What number comes before {n}?", n - 1
        if 'one more' in skill.skill_text.lower():
            return f"What is one more than {n}?", n + 1
        return f"What number comes after {n}?", n + 1
    if generator_type == 'number_bonds':
        target = int(config.get('target') or (5 if skill.level <= 1 else 10))
        a = random.randint(0, target)
        if skill.strand.name == 'Number Sense' and target in {5, 10}:
            visual_type = 'five_frame' if target == 5 else 'ten_frame'
            return f"How many more dots fill the frame? {_visual_marker(visual_type, a)}", target - a
        return f"What goes with {a} to make {target}?", target - a
    if generator_type == 'addition':
        variables = _generate_variables(template.generator_config_json)
        a, b = variables.get('a', 1), variables.get('b', 1)
        return f"What is {a} + {b}?", a + b
    if generator_type == 'subtraction':
        variables = _generate_variables(template.generator_config_json)
        a, b = max(variables.get('a', 1), variables.get('b', 1)), min(variables.get('a', 1), variables.get('b', 1))
        return f"What is {a} - {b}?", a - b
    if generator_type == 'multiplication':
        variables = _generate_variables(template.generator_config_json)
        a, b = variables.get('a', 1), variables.get('b', 1)
        return f"What is {a} × {b}?", a * b
    if generator_type == 'division':
        variables = _generate_variables(template.generator_config_json)
        divisor = max(1, variables.get('b', 1))
        quotient = max(1, variables.get('a', 1))
        dividend = divisor * quotient
        return f"What is {dividend} ÷ {divisor}?", quotient

    variables = _generate_variables(template.generator_config_json)
    question_text = prompt
    for name, value in variables.items():
        question_text = question_text.replace('{' + name + '}', str(value))
    correct_answer = _safe_answer(question_text, variables)
    return question_text, correct_answer



def _visual_marker(visual_type: str, value: int) -> str:
    return f'[[mf_visual:{visual_type}:{max(0, min(10, int(value)))}]]'


def question_display_text(question: MathsFundamentalQuestion | None) -> str:
    if not question:
        return ''
    return VARIANT_MARKER_RE.sub('', VISUAL_MARKER_RE.sub('', question.question_text or '')).strip()


def question_visual_model(question: MathsFundamentalQuestion | None) -> dict[str, Any] | None:
    if not question or not question.question_text:
        return None
    match = VISUAL_MARKER_RE.search(question.question_text)
    if not match:
        return None
    visual_type, raw_value = match.groups()
    value = max(0, min(10, int(raw_value)))
    cells = [index < value for index in range(10 if visual_type == 'ten_frame' else 5)]
    if visual_type == 'dot_pattern':
        cells = [index < value for index in range(value)]
    return {'type': visual_type, 'value': value, 'cells': cells}


def level_progress(attempt: MathsFundamentalAttempt) -> dict[str, Any]:
    total = QUESTIONS_PER_LEVEL
    answered = MathsFundamentalQuestion.query.filter_by(attempt_id=attempt.id, level=attempt.current_level).filter(MathsFundamentalQuestion.pupil_answer.isnot(None)).count()
    return {'answered': min(answered, total), 'total': total, 'remaining': max(0, total - answered)}


def attempt_completion_summary(attempt: MathsFundamentalAttempt) -> dict[str, Any]:
    secure_level = attempt.final_level if attempt.final_level is not None else max(0, attempt.current_level - 1)
    result = MathsFundamentalResult.query.filter_by(
        school_id=attempt.school_id,
        pupil_id=attempt.pupil_id,
        academic_year=attempt.academic_year,
        strand_id=attempt.strand_id,
    ).first()
    if result:
        secure_level = max(int(result.current_level or 0), int(secure_level or 0))
    next_level = min(MAX_LEVEL, secure_level + 1)
    next_skill = MathsFundamentalSkill.query.filter_by(strand_id=attempt.strand_id, level=next_level).order_by(MathsFundamentalSkill.display_order, MathsFundamentalSkill.id).first()
    return {
        'secure_level': secure_level,
        'next_level': next_level if next_skill else None,
        'teaching_focus': (next_skill.teaching_prompt or next_skill.skill_text) if next_skill else 'Maintain and deepen secure learning.',
    }

def choose_prompt(prompt_text: str) -> str:
    prompts = [part.strip() for part in re.split(r';|\n', prompt_text or '') if part.strip()]
    return random.choice(prompts) if prompts else 'Show what you know about this skill.'


def _load_generator_config(config_text: str | None) -> dict[str, Any]:
    try:
        config = json.loads(config_text or '{}')
    except json.JSONDecodeError:
        return {}
    return config if isinstance(config, dict) else {}


def _bounds(value: Any, default_low: int, default_high: int) -> tuple[int, int]:
    if isinstance(value, list) and len(value) >= 2:
        return int(value[0]), int(value[1])
    return default_low, default_high


def _generate_variables(config_text: str | None) -> dict[str, int]:
    try:
        config = json.loads(config_text or '{}')
    except json.JSONDecodeError:
        config = {}
    variables = {}
    for key, bounds in config.items():
        if isinstance(bounds, list) and len(bounds) >= 2:
            variables[key] = random.randint(int(bounds[0]), int(bounds[1]))
        elif isinstance(bounds, int):
            variables[key] = random.randint(1, bounds)
    if not variables:
        variables = {'a': random.randint(1, 12), 'b': random.randint(1, 12)}
    return variables


def _safe_answer(template_text: str, variables: dict[str, int]) -> int | float | None:
    expression = template_text.lower().strip().rstrip('?')
    expression = expression.replace('what is', '').replace('calculate', '').replace('×', '*').replace('x', '*').replace('÷', '/')
    for key, value in variables.items():
        expression = expression.replace('{' + key + '}', str(value))
    expression = expression.strip()
    if not re.fullmatch(r'[0-9+\-*/ ().]+', expression):
        return None
    try:
        value = eval(expression, {'__builtins__': {}}, {})  # noqa: S307 - expression is character-whitelisted above.
    except Exception:
        return None
    return int(value) if float(value).is_integer() else round(float(value), 2)


def start_session(*, school_id: int, teacher_id: int, class_id: int | None, strand_id: int, academic_year: str, starting_level: int, questions_per_level: int, group_name: str | None = None) -> MathsFundamentalsSession:
    open_sessions = MathsFundamentalsSession.query.filter_by(school_id=school_id, teacher_id=teacher_id, is_open=True)
    if class_id:
        open_sessions = open_sessions.filter_by(class_id=class_id)
    for session in open_sessions.all():
        session.is_open = False
        session.closed_at = datetime.now(timezone.utc)
        db.session.add(session)
    session = MathsFundamentalsSession(
        school_id=school_id,
        teacher_id=teacher_id,
        class_id=class_id,
        strand_id=strand_id,
        academic_year=academic_year,
        starting_level=max(1, min(15, starting_level)),
        questions_per_level=QUESTIONS_PER_LEVEL,
        group_name=group_name,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=3),
    )
    db.session.add(session)
    db.session.commit()
    return session


def close_session(session: MathsFundamentalsSession) -> None:
    session.is_open = False
    session.closed_at = datetime.now(timezone.utc)
    db.session.add(session)
    db.session.commit()


def active_session_for_pupil(pupil: Pupil) -> MathsFundamentalsSession | None:
    now = datetime.now(timezone.utc)
    return MathsFundamentalsSession.query.filter(
        MathsFundamentalsSession.school_id == pupil.school_id,
        MathsFundamentalsSession.class_id == pupil.class_id,
        MathsFundamentalsSession.is_open.is_(True),
        (MathsFundamentalsSession.expires_at.is_(None) | (MathsFundamentalsSession.expires_at > now)),
    ).order_by(MathsFundamentalsSession.opened_at.desc()).first()


def get_or_start_attempt(session: MathsFundamentalsSession, pupil: Pupil) -> MathsFundamentalAttempt:
    attempt = MathsFundamentalAttempt.query.filter_by(session_id=session.id, pupil_id=pupil.id).order_by(MathsFundamentalAttempt.started_at.desc()).first()
    if attempt and attempt.status == 'in_progress':
        if attempt.questions_per_level != QUESTIONS_PER_LEVEL:
            attempt.questions_per_level = QUESTIONS_PER_LEVEL
            db.session.add(attempt)
            db.session.commit()
        return attempt
    attempt = MathsFundamentalAttempt(
        school_id=session.school_id,
        pupil_id=pupil.id,
        strand_id=session.strand_id,
        academic_year=session.academic_year,
        session_id=session.id,
        current_level=session.starting_level,
        questions_per_level=QUESTIONS_PER_LEVEL,
        last_activity_at=datetime.now(timezone.utc),
    )
    db.session.add(attempt)
    db.session.commit()
    return attempt


def next_question_for_attempt(attempt: MathsFundamentalAttempt) -> MathsFundamentalQuestion | None:
    unanswered = MathsFundamentalQuestion.query.filter_by(attempt_id=attempt.id, pupil_answer=None).order_by(MathsFundamentalQuestion.id).first()
    if unanswered:
        return unanswered
    if attempt.status == 'completed':
        return None
    if attempt.questions_per_level != QUESTIONS_PER_LEVEL:
        attempt.questions_per_level = QUESTIONS_PER_LEVEL
    answered_level = MathsFundamentalQuestion.query.filter_by(attempt_id=attempt.id, level=attempt.current_level).filter(MathsFundamentalQuestion.pupil_answer.isnot(None)).count()
    if answered_level >= QUESTIONS_PER_LEVEL:
        return None
    skill = MathsFundamentalSkill.query.filter_by(strand_id=attempt.strand_id, level=attempt.current_level).order_by(MathsFundamentalSkill.display_order, MathsFundamentalSkill.id).first()
    if not skill:
        complete_attempt(attempt, max(0, attempt.current_level - 1))
        return None
    question = generate_unique_question(skill, attempt)
    question.attempt_id = attempt.id
    db.session.add(question)
    attempt.last_activity_at = datetime.now(timezone.utc)
    db.session.add(attempt)
    db.session.commit()
    return question



def generate_unique_question(skill: MathsFundamentalSkill, attempt: MathsFundamentalAttempt) -> MathsFundamentalQuestion:
    previous_texts = {
        VARIANT_MARKER_RE.sub('', row.question_text or '').strip()
        for row in MathsFundamentalQuestion.query.join(MathsFundamentalAttempt).filter(
            MathsFundamentalAttempt.pupil_id == attempt.pupil_id,
        ).all()
    }
    question = generate_question(skill, pupil_id=attempt.pupil_id)
    for _ in range(12):
        if VARIANT_MARKER_RE.sub('', question.question_text or '').strip() not in previous_texts:
            return question
        question = generate_question(skill, pupil_id=attempt.pupil_id)
    suffix = random.randint(1000, 9999)
    question.question_text = f'{question.question_text} [[mf_variant:{suffix}]]'
    return question

def submit_answer(question: MathsFundamentalQuestion, answer: str) -> MathsFundamentalAttempt:
    question.pupil_answer = (answer or '').strip()
    if question.teacher_mark_required:
        question.is_correct = None
    else:
        question.is_correct = normalize_answer(question.pupil_answer) == normalize_answer(question.correct_answer)
    question.answered_at = datetime.now(timezone.utc)
    attempt = question.attempt
    attempt.last_activity_at = datetime.now(timezone.utc)
    db.session.add(question)
    db.session.add(attempt)
    db.session.flush()
    evaluate_attempt_progress(attempt)
    db.session.commit()
    return attempt


def normalize_answer(value: str | None) -> str:
    value = (value or '').strip().lower()
    try:
        numeric = float(value)
        return str(int(numeric)) if numeric.is_integer() else str(round(numeric, 2))
    except ValueError:
        return re.sub(r'\s+', ' ', value)


def evaluate_attempt_progress(attempt: MathsFundamentalAttempt) -> None:
    if attempt.questions_per_level != QUESTIONS_PER_LEVEL:
        attempt.questions_per_level = QUESTIONS_PER_LEVEL
    questions = MathsFundamentalQuestion.query.filter_by(attempt_id=attempt.id, level=attempt.current_level).filter(MathsFundamentalQuestion.pupil_answer.isnot(None)).all()
    if len(questions) < QUESTIONS_PER_LEVEL or any(q.is_correct is None for q in questions):
        return
    correct = sum(1 for q in questions if q.is_correct)
    if correct >= PASS_MARK:
        next_level = attempt.current_level + 1
        if next_level > MAX_LEVEL or not MathsFundamentalSkill.query.filter_by(strand_id=attempt.strand_id, level=next_level).first():
            complete_attempt(attempt, attempt.current_level)
        else:
            attempt.current_level = next_level
            db.session.add(attempt)
    else:
        complete_attempt(attempt, max(0, attempt.current_level - 1))


def complete_attempt(attempt: MathsFundamentalAttempt, final_level: int) -> None:
    attempt.status = 'completed'
    attempt.completed_at = datetime.now(timezone.utc)
    attempt.final_level = final_level
    attempt.last_activity_at = datetime.now(timezone.utc)
    result = MathsFundamentalResult.query.filter_by(
        school_id=attempt.school_id,
        pupil_id=attempt.pupil_id,
        academic_year=attempt.academic_year,
        strand_id=attempt.strand_id,
    ).first()
    if not result:
        result = MathsFundamentalResult(school_id=attempt.school_id, pupil_id=attempt.pupil_id, academic_year=attempt.academic_year, strand_id=attempt.strand_id)
    awarded_level = max(int(result.current_level or 0), int(final_level or 0))
    skill = MathsFundamentalSkill.query.filter_by(strand_id=attempt.strand_id, level=max(1, awarded_level)).order_by(MathsFundamentalSkill.display_order).first()
    next_skill = MathsFundamentalSkill.query.filter_by(strand_id=attempt.strand_id, level=min(MAX_LEVEL, awarded_level + 1)).order_by(MathsFundamentalSkill.display_order).first()
    result.current_level = awarded_level
    result.current_skill_id = skill.id if skill else None
    result.last_assessed = attempt.completed_at
    result.next_step = (next_skill.teaching_prompt or next_skill.skill_text) if next_skill else 'Maintain and deepen secure learning.'
    db.session.add(attempt)
    db.session.add(result)


def attempt_status_rows(session: MathsFundamentalsSession) -> list[dict[str, Any]]:
    pupils = Pupil.query.filter_by(class_id=session.class_id, school_id=session.school_id, is_active=True).order_by(Pupil.last_name, Pupil.first_name).all() if session.class_id else []
    attempts = {}
    for attempt in sorted(session.attempts, key=lambda row: row.id or 0):
        attempts[attempt.pupil_id] = attempt
    rows = []
    for pupil in pupils:
        attempt = attempts.get(pupil.id)
        current_question = None
        if attempt:
            current_question = MathsFundamentalQuestion.query.filter_by(attempt_id=attempt.id, pupil_answer=None).order_by(MathsFundamentalQuestion.id).first()
        rows.append({
            'pupil': pupil,
            'attempt': attempt,
            'status': 'Completed' if attempt and attempt.status == 'completed' else ('In Progress' if attempt and attempt.status == 'in_progress' else ('Restarted' if attempt else 'Waiting')),
            'current_question': question_display_text(current_question) if current_question else '—',
            'current_skill': current_question.skill.skill_text if current_question and current_question.skill else '—',
            'current_level': attempt.current_level if attempt else session.starting_level,
            'last_activity': attempt.last_activity_at if attempt else None,
        })
    return rows


def class_and_admin_summary(pupils: list[Pupil], strands: list[MathsFundamentalStrand], academic_year: str) -> dict[str, Any]:
    results = current_results_map(pupils, academic_year)
    by_strand = {}
    for strand in strands:
        values = [result.current_level for (pupil_id, strand_id), result in results.items() if strand_id == strand.id]
        by_strand[strand.id] = round(mean(values), 1) if values else None
    class_values = defaultdict(list)
    year_values = defaultdict(list)
    for pupil in pupils:
        for strand in strands:
            result = results.get((pupil.id, strand.id))
            if result:
                class_values[pupil.school_class.name].append(result.current_level)
                year_values[pupil.school_class.year_group].append(result.current_level)
    return {
        'average_by_strand': by_strand,
        'average_by_class': {name: round(mean(values), 1) for name, values in class_values.items() if values},
        'average_by_year': {year: round(mean(values), 1) for year, values in year_values.items() if values},
    }


def intervention_candidates(pupils: list[Pupil], strands: list[MathsFundamentalStrand], academic_year: str) -> list[dict[str, Any]]:
    pupil_ids = [p.id for p in pupils]
    if not pupil_ids:
        return []
    latest = MathsFundamentalResult.query.filter(MathsFundamentalResult.pupil_id.in_(pupil_ids), MathsFundamentalResult.academic_year == academic_year).all()
    pupils_by_id = {p.id: p for p in pupils}
    strands_by_id = {s.id: s for s in strands}
    candidates = []
    for result in latest:
        if result.current_level <= 4:
            candidates.append({
                'pupil': pupils_by_id.get(result.pupil_id),
                'strand': strands_by_id.get(result.strand_id),
                'current_level': result.current_level,
                'suggested_teaching': result.next_step or 'Revisit prerequisite skills with concrete resources.',
                'reason': 'Current level is 4 or below.',
            })
            continue
        attempts = MathsFundamentalAttempt.query.filter_by(pupil_id=result.pupil_id, strand_id=result.strand_id, academic_year=academic_year, status='completed').order_by(MathsFundamentalAttempt.completed_at.desc()).limit(2).all()
        if len(attempts) == 2 and attempts[0].final_level is not None and attempts[1].final_level is not None and attempts[0].final_level <= attempts[1].final_level:
            candidates.append({
                'pupil': pupils_by_id.get(result.pupil_id),
                'strand': strands_by_id.get(result.strand_id),
                'current_level': result.current_level,
                'suggested_teaching': result.next_step or 'Target the next small step and reassess after practice.',
                'reason': 'Two assessments without progress.',
            })
    return [row for row in candidates if row['pupil'] and row['strand']]


def filtered_school_pupils(school_id: int | None, filters: dict[str, str], class_id: str | None = None, year_group: str | None = None):
    query = Pupil.query.join(SchoolClass).filter(Pupil.is_active.is_(True))
    if school_id is not None:
        query = query.filter(Pupil.school_id == school_id)
    if class_id:
        query = query.filter(Pupil.class_id == int(class_id))
    if year_group not in (None, ''):
        query = query.filter(SchoolClass.year_group == int(year_group))
    query = apply_admin_pupil_filters(query, filters)
    return query.order_by(SchoolClass.year_group, SchoolClass.name, Pupil.last_name, Pupil.first_name).all()
